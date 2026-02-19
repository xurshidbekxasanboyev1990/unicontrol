import openpyxl
import glob
import os

files = glob.glob(os.path.join(r"c:\Users\user\Desktop\unicontrol", "1-2*dars jadvali*.xlsx"))
if not files:
    print("File not found!")
    exit()

fname = files[0]
print(f"Reading: {fname}")
wb = openpyxl.load_workbook(fname)
print('Sheets:', wb.sheetnames)
print()

# Find the last sheet or teacher-related sheet
last_sheet = wb.sheetnames[-1]
print(f"Last sheet: {last_sheet}")

ws = wb[last_sheet]
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print(f"Merged cells: {len(list(ws.merged_cells.ranges))}")
print()

# Print ALL rows to see full structure
for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=False):
    vals = []
    for c in row:
        if c.value is not None:
            vals.append(f"[{c.column_letter}{c.row}]={c.value}")
    if vals:
        print(f"Row {row[0].row}: " + " | ".join(vals))
