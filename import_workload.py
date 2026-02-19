"""
Import teacher workload from Excel to UniControl database.
Parses the 'Oqituvchi bandligi' sheet and imports via API.

Usage: py import_workload.py
"""

import openpyxl
import glob
import os
import json
import requests
import sys

BASE_URL = "http://localhost:8001/api/v1"

# === Parse Excel ===
def parse_excel():
    files = glob.glob(os.path.join(os.path.dirname(__file__), "1-2*dars jadvali*.xlsx"))
    if not files:
        print("Excel fayl topilmadi!")
        return []
    
    fname = files[0]
    print(f"Excel o'qilmoqda: {fname}")
    wb = openpyxl.load_workbook(fname)
    
    if "Oqituvchi bandligi" not in wb.sheetnames:
        print("'Oqituvchi bandligi' sheet topilmadi!")
        print(f"Mavjud sheetlar: {wb.sheetnames}")
        return []
    
    ws = wb["Oqituvchi bandligi"]
    
    # Build merged cells map
    merged_map = {}
    for mc in ws.merged_cells.ranges:
        for row in range(mc.min_row, mc.max_row + 1):
            for col in range(mc.min_col, mc.max_col + 1):
                merged_map[(row, col)] = ws.cell(mc.min_row, mc.min_col).value
    
    def get_val(row, col):
        v = ws.cell(row, col).value
        if v is not None:
            return v
        return merged_map.get((row, col), None)
    
    # Parse department ranges from row 1
    dept_ranges = {}
    for col in range(4, ws.max_column + 1):
        v = get_val(1, col)
        if v:
            dept_name = str(v).strip()
            if dept_name not in dept_ranges:
                dept_ranges[dept_name] = {"start": col, "end": col}
            dept_ranges[dept_name]["end"] = col
    
    # Parse teachers from row 2
    teachers = {}
    for col in range(4, ws.max_column + 1):
        v = get_val(2, col)
        if v:
            teacher_type = get_val(5, col) or ""
            dept = ""
            for d_name, d_range in dept_ranges.items():
                if d_range["start"] <= col <= d_range["end"]:
                    dept = d_name
                    break
            teachers[col] = {
                "name": str(v).strip(),
                "type": str(teacher_type).strip(),
                "department": dept
            }
    
    # Day/time mappings
    DAYS_MAP = {
        "Dushanba": ("monday", "Dushanba"),
        "Seshanba": ("tuesday", "Seshanba"),
        "Chorshanba": ("wednesday", "Chorshanba"),
        "Payshanba": ("thursday", "Payshanba"),
        "Juma": ("friday", "Juma"),
        "Juma ": ("friday", "Juma"),
        "Shanba": ("saturday", "Shanba"),
    }
    
    TIME_MAP = {
        1: ("08:30", "09:50"),
        2: ("10:00", "11:20"),
        3: ("12:00", "13:20"),
        4: ("13:30", "14:50"),
        5: ("15:00", "16:20"),
        6: ("16:30", "17:50"),
    }
    
    # Parse schedule entries
    entries = []
    current_day = None
    current_day_en = None
    current_day_uz = None
    
    for row in range(6, ws.max_row + 1):
        day_val = get_val(row, 1)
        lesson_num = get_val(row, 2)
        
        if day_val:
            day_str = str(day_val).strip()
            if day_str in DAYS_MAP:
                current_day_en, current_day_uz = DAYS_MAP[day_str]
        
        if not lesson_num or not current_day_en:
            continue
        
        try:
            ln = int(float(lesson_num))
        except (ValueError, TypeError):
            continue
        
        start_t, end_t = TIME_MAP.get(ln, ("", ""))
        
        for col, teacher in teachers.items():
            group_val = get_val(row, col)
            if group_val and str(group_val).strip():
                gv = str(group_val).strip()
                entries.append({
                    "teacher_name": teacher["name"],
                    "department": teacher["department"],
                    "teacher_type": teacher["type"],
                    "day_of_week": current_day_en,
                    "day_name_uz": current_day_uz,
                    "lesson_number": ln,
                    "start_time": start_t,
                    "end_time": end_t,
                    "groups": gv,
                    "is_busy": gv == "BAND"
                })
    
    print(f"Kafedralar: {len(dept_ranges)}")
    print(f"O'qituvchilar: {len(teachers)}")
    print(f"Dars yozuvlari: {len(entries)}")
    
    return entries


# === Login and Import ===
def login(username="superadmin", password="superadmin123"):
    try:
        r = requests.post(f"{BASE_URL}/auth/login", json={"login": username, "password": password})
        r.raise_for_status()
        return r.json()["access_token"]
    except Exception as e:
        print(f"Login xatosi: {e}")
        return None


def import_to_api(entries, token):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    
    # Import in batches of 500
    batch_size = 500
    total_imported = 0
    
    for i in range(0, len(entries), batch_size):
        batch = entries[i:i+batch_size]
        clear = (i == 0)  # Clear only on first batch
        
        payload = {
            "items": batch,
            "clear_existing": clear
        }
        
        try:
            r = requests.post(
                f"{BASE_URL}/academic/workload/import",
                json=payload,
                headers=headers,
                timeout=30
            )
            r.raise_for_status()
            result = r.json()
            total_imported += result.get("total_imported", 0)
            print(f"  Batch {i//batch_size + 1}: {len(batch)} ta yuklandi")
        except Exception as e:
            print(f"  Batch {i//batch_size + 1} xatosi: {e}")
            if hasattr(e, 'response') and e.response:
                print(f"  Response: {e.response.text}")
    
    return total_imported


def main():
    print("=" * 50)
    print("O'qituvchi bandligi import qilish")
    print("=" * 50)
    
    # Parse Excel
    entries = parse_excel()
    if not entries:
        print("Ma'lumot topilmadi!")
        return
    
    # Login
    print("\nTizimga kirish...")
    token = login()
    if not token:
        print("Login muvaffaqiyatsiz!")
        return
    print("  OK")
    
    # Import
    print("\nImport qilish...")
    total = import_to_api(entries, token)
    print(f"\nJami: {total} ta yozuv import qilindi!")
    print("=" * 50)


if __name__ == "__main__":
    main()
