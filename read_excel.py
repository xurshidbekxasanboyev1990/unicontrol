import openpyxl
import json
import glob
import os

# Find the xlsx file
files = glob.glob(os.path.join(r"c:\Users\user\Desktop\unicontrol", "1-2*dars jadvali*.xlsx"))
if not files:
    print("File not found!")
    exit()

fname = files[0]
print(f"Reading: {fname}")
wb = openpyxl.load_workbook(fname)
print('Sheet names:', wb.sheetnames)
print(f'Total sheets: {len(wb.sheetnames)}')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    merged = list(ws.merged_cells.ranges)
    print(f'=== Sheet: {sheet_name} === (rows={ws.max_row}, cols={ws.max_column}, merged={len(merged)})')
    if merged:
        print(f'  Sample merges: {[str(m) for m in merged[:5]]}')
    # Print first 6 rows to see structure
    for row in ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=False):
        vals = []
        for c in row:
            if c.value is not None:
                vals.append(f"[{c.column_letter}{c.row}]={c.value}")
        if vals:
            print(f'  Row {row[0].row}: {", ".join(vals[:6])}{"..." if len(vals) > 6 else ""}')
    print()
