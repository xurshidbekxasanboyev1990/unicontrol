import openpyxl
import glob
import os
import json

files = glob.glob(os.path.join(r"c:\Users\user\Desktop\unicontrol", "1-2*dars jadvali*.xlsx"))
fname = files[0]
wb = openpyxl.load_workbook(fname)

ws = wb["Oqituvchi bandligi"]

# Get merged cells info
merged_map = {}
for mc in ws.merged_cells.ranges:
    for row in range(mc.min_row, mc.max_row + 1):
        for col in range(mc.min_col, mc.max_col + 1):
            merged_map[(row, col)] = ws.cell(mc.min_row, mc.min_col).value

def get_val(row, col):
    v = ws.cell(row, col).value
    if v is not None:
        return v
    return merged_map.get((row, col), None)

# Structure:
# Row 1: Header with department names (kafedra)
# Row 2: Teacher names (one per column)
# Row 3-4: ???
# Row 5: Teacher type (Asosiy, Tashqi, etc.)
# Rows 6+: Schedule - col A=day, col B=lesson#, col C=time, col D+=group names

# Let's parse the structure properly
# First, find all departments and their column ranges from row 1
departments = {}
for col in range(1, ws.max_column + 1):
    v = get_val(1, col)
    if v and col >= 4:  # Skip A-C (day/lesson/time columns)
        departments[col] = str(v)

print("=== DEPARTMENTS ===")
for col, name in sorted(departments.items()):
    letter = openpyxl.utils.get_column_letter(col)
    print(f"  Col {letter} ({col}): {name}")

# Parse teachers from row 2
teachers = {}
for col in range(4, ws.max_column + 1):
    v = get_val(2, col)
    if v:
        letter = openpyxl.utils.get_column_letter(col)
        teacher_type = get_val(5, col) or "N/A"
        teachers[col] = {"name": str(v).strip(), "type": str(teacher_type).strip(), "col_letter": letter}

print(f"\n=== TEACHERS ({len(teachers)} total) ===")
for col, t in sorted(teachers.items()):
    print(f"  Col {t['col_letter']}: {t['name']} ({t['type']})")

# Parse schedule rows
print("\n=== SCHEDULE STRUCTURE ===")
days_schedule = {}
current_day = None
for row in range(6, ws.max_row + 1):
    day_val = get_val(row, 1)  # Col A - day name
    lesson_num = get_val(row, 2)  # Col B - lesson number
    time_val = get_val(row, 3)  # Col C - time
    
    if day_val:
        current_day = str(day_val).strip()
    
    if lesson_num and current_day:
        # For each teacher column, check what group they're teaching
        for col, teacher in teachers.items():
            group_val = get_val(row, col)
            if group_val and str(group_val).strip() and str(group_val).strip() != "BAND":
                key = f"{current_day}|{int(lesson_num)}|{teacher['name']}"
                if key not in days_schedule:
                    days_schedule[key] = {
                        "day": current_day,
                        "lesson_number": int(lesson_num),
                        "time": str(time_val) if time_val else "",
                        "teacher_name": teacher['name'],
                        "teacher_type": teacher['type'],
                        "groups": str(group_val).strip()
                    }

print(f"\nTotal schedule entries: {len(days_schedule)}")

# Print first 20 entries as sample
for i, (key, entry) in enumerate(sorted(days_schedule.items())[:20]):
    print(f"  {entry['day']} | Para {entry['lesson_number']} | {entry['teacher_name']} -> {entry['groups']}")

# Also check if there's a second table in the sheet (rows 46+)
print("\n=== ROWS 46-70 (secondary table?) ===")
for row in range(46, min(70, ws.max_row + 1)):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = get_val(row, col)
        if v:
            letter = openpyxl.utils.get_column_letter(col)
            vals.append(f"[{letter}{row}]={v}")
    if vals:
        print(f"  Row {row}: {' | '.join(vals[:15])}")
