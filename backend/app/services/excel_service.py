"""
UniControl - Universal Excel Service v2
========================================
Handles ALL Excel import/export operations with UPSERT logic.
- Kontingent import (students + users + groups) — UPSERT
- Schedule import from Excel — UPSERT with fuzzy group matching
- Attendance import — UPSERT
- Students / Groups import — UPSERT
- All exports (students, groups, attendance, payments, schedules, reports)

Key principles:
1. If record exists → UPDATE it
2. If record is new → INSERT it
3. If record was in DB but NOT in new file → optionally deactivate
4. Fuzzy group matching to handle naming differences (KI-25-09 vs KI_25-09)

Author: UniControl Team
Version: 2.0.0
"""

import io
import re
import logging
from datetime import datetime, date, time as dt_time
from decimal import Decimal
from typing import Optional, List, Dict, Any, Set
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from sqlalchemy.orm import joinedload

from app.config import now_tashkent
from app.models.student import Student
from app.models.group import Group
from app.models.attendance import Attendance, AttendanceStatus
from app.models.schedule import Schedule, WeekDay, ScheduleType
from app.core.exceptions import BadRequestException

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# HELPER: Fuzzy Group Name Matching
# ═══════════════════════════════════════════════════════════════

def normalize_group_name(name: str) -> str:
    """Normalize group name for comparison: strip, upper, replace separators."""
    n = name.strip().upper()
    # Remove prefixes like "Magistr_", "Bakalavr_"
    n = re.sub(r'^(MAGISTR|BAKALAVR|MAGISTRATURA|BAKALAVRIATURA)[_\s\-]+', '', n, flags=re.IGNORECASE)
    n = re.sub(r'[\s_\-]+', '-', n)   # unify separators to '-'
    return n.strip('-').strip()


def fuzzy_match_group(
    name: str,
    candidates: Dict[str, Any],
    threshold: float = 0.70
) -> Optional[str]:
    """
    Find best matching group name from candidates dict.
    Returns the matched candidate key or None.
    
    Strategy:
    1. Exact match (after normalization)
    2. SequenceMatcher ratio >= threshold
    """
    norm = normalize_group_name(name)

    # Build normalized lookup
    norm_to_orig = {}
    for c in candidates:
        nc = normalize_group_name(c)
        norm_to_orig[nc] = c

    # Exact normalized match
    if norm in norm_to_orig:
        return norm_to_orig[norm]

    # Fuzzy match
    best_match = None
    best_score = 0.0
    for nc, orig in norm_to_orig.items():
        score = SequenceMatcher(None, norm, nc).ratio()
        if score > best_score and score >= threshold:
            best_score = score
            best_match = orig

    return best_match


def build_group_lookup(groups: list) -> Dict[str, Any]:
    """Build a lookup dict: name → group object, with normalized variants."""
    lookup = {}
    for g in groups:
        lookup[g.name] = g
        lookup[g.name.upper()] = g
        lookup[normalize_group_name(g.name)] = g
    return lookup


# ═══════════════════════════════════════════════════════════════
# HELPER: Day / Time Parsing
# ═══════════════════════════════════════════════════════════════

DAY_MAP = {
    # Uzbek
    "dushanba": WeekDay.MONDAY, "du": WeekDay.MONDAY,
    "seshanba": WeekDay.TUESDAY, "se": WeekDay.TUESDAY,
    "chorshanba": WeekDay.WEDNESDAY, "cho": WeekDay.WEDNESDAY, "chor": WeekDay.WEDNESDAY,
    "payshanba": WeekDay.THURSDAY, "pa": WeekDay.THURSDAY, "pay": WeekDay.THURSDAY,
    "juma": WeekDay.FRIDAY, "ju": WeekDay.FRIDAY,
    "shanba": WeekDay.SATURDAY, "sha": WeekDay.SATURDAY,
    "yakshanba": WeekDay.SUNDAY, "yak": WeekDay.SUNDAY,
    # Russian
    "понедельник": WeekDay.MONDAY, "пн": WeekDay.MONDAY,
    "вторник": WeekDay.TUESDAY, "вт": WeekDay.TUESDAY,
    "среда": WeekDay.WEDNESDAY, "ср": WeekDay.WEDNESDAY,
    "четверг": WeekDay.THURSDAY, "чт": WeekDay.THURSDAY,
    "пятница": WeekDay.FRIDAY, "пт": WeekDay.FRIDAY,
    "суббота": WeekDay.SATURDAY, "сб": WeekDay.SATURDAY,
    "воскресенье": WeekDay.SUNDAY, "вс": WeekDay.SUNDAY,
    # English
    "monday": WeekDay.MONDAY, "mon": WeekDay.MONDAY,
    "tuesday": WeekDay.TUESDAY, "tue": WeekDay.TUESDAY,
    "wednesday": WeekDay.WEDNESDAY, "wed": WeekDay.WEDNESDAY,
    "thursday": WeekDay.THURSDAY, "thu": WeekDay.THURSDAY,
    "friday": WeekDay.FRIDAY, "fri": WeekDay.FRIDAY,
    "saturday": WeekDay.SATURDAY, "sat": WeekDay.SATURDAY,
    "sunday": WeekDay.SUNDAY, "sun": WeekDay.SUNDAY,
    # Numbers
    "1": WeekDay.MONDAY, "2": WeekDay.TUESDAY, "3": WeekDay.WEDNESDAY,
    "4": WeekDay.THURSDAY, "5": WeekDay.FRIDAY, "6": WeekDay.SATURDAY,
    "7": WeekDay.SUNDAY,
}

# Default lesson time slots (standard schedule)
DEFAULT_TIME_SLOTS = {
    1: ("08:30", "09:50"),
    2: ("10:00", "11:20"),
    3: ("12:00", "13:20"),
    4: ("13:30", "14:50"),
    5: ("15:00", "16:20"),
    6: ("16:30", "17:50"),
    7: ("18:00", "19:20"),
}


def parse_day(raw: Any) -> Optional[WeekDay]:
    """Parse a day value (string/int) to WeekDay enum."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    return DAY_MAP.get(s)


def parse_time(raw: Any) -> Optional[dt_time]:
    """Parse time from various formats."""
    if raw is None:
        return None
    if isinstance(raw, dt_time):
        return raw
    if isinstance(raw, datetime):
        return raw.time()
    s = str(raw).strip()
    for fmt in ["%H:%M", "%H:%M:%S", "%H.%M", "%I:%M %p"]:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    # Try simple split
    m = re.match(r'(\d{1,2})[:\.\-](\d{2})', s)
    if m:
        return dt_time(int(m.group(1)), int(m.group(2)))
    return None


# ═══════════════════════════════════════════════════════════════
# MAIN SERVICE
# ═══════════════════════════════════════════════════════════════

class ExcelService:
    """Universal Excel import/export service with UPSERT logic."""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ══════════════════════════════════════════════════════
    # EXPORT METHODS
    # ══════════════════════════════════════════════════════

    async def export_students(
        self,
        group_id: Optional[int] = None,
        include_all_columns: bool = False
    ) -> io.BytesIO:
        """Export students to Excel."""
        query = select(Student).options(joinedload(Student.group))
        if group_id:
            query = query.where(Student.group_id == group_id)
        query = query.order_by(Student.name)

        result = await self.db.execute(query)
        students = result.unique().scalars().all()

        data = []
        for s in students:
            row = {
                "ID": s.student_id,
                "F.I.O": s.name,
                "Guruh": s.group.name if s.group else "",
                "Telefon": s.phone or "",
                "Email": s.email or "",
                "Tug'ilgan sana": s.birth_date.strftime("%d.%m.%Y") if s.birth_date else "",
                "Jinsi": "Erkak" if s.gender == "male" else "Ayol" if s.gender == "female" else "",
                "Kontrakt": float(s.contract_amount),
                "To'langan": float(s.contract_paid),
                "Qoldi": float(s.contract_remaining),
                "Holati": "Faol" if s.is_active else "Nofaol",
            }
            if include_all_columns:
                row.update({
                    "Manzil": s.address or "",
                    "Transport": s.commute or "",
                    "Pasport": s.passport or "",
                    "JSHSHIR": s.jshshir or "",
                    "Qabul sanasi": s.enrollment_date.strftime("%d.%m.%Y") if s.enrollment_date else "",
                    "Bitirish sanasi": s.graduation_date.strftime("%d.%m.%Y") if s.graduation_date else "",
                    "Guruh lideri": "Ha" if s.is_leader else "Yo'q",
                })
            data.append(row)

        return self._create_excel_file(pd.DataFrame(data), "Talabalar ro'yxati")

    async def export_attendance(
        self,
        group_id: Optional[int] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None
    ) -> io.BytesIO:
        """Export attendance to Excel."""
        query = select(Attendance).options(
            joinedload(Attendance.student).joinedload(Student.group)
        )
        if group_id:
            query = query.join(Student).where(Student.group_id == group_id)
        if date_from:
            query = query.where(Attendance.date >= date_from)
        if date_to:
            query = query.where(Attendance.date <= date_to)
        query = query.order_by(Attendance.date.desc(), Attendance.student_id)

        result = await self.db.execute(query)
        attendances = result.unique().scalars().all()

        status_map = {
            AttendanceStatus.PRESENT: "Keldi",
            AttendanceStatus.ABSENT: "Kelmadi",
            AttendanceStatus.LATE: "Kech qoldi",
            AttendanceStatus.EXCUSED: "Sababli",
        }
        data = []
        for a in attendances:
            data.append({
                "Sana": a.date.strftime("%d.%m.%Y"),
                "Talaba": a.student.name if a.student else "",
                "Guruh": a.student.group.name if a.student and a.student.group else "",
                "Holat": status_map.get(a.status, ""),
                "Kechikish (min)": a.late_minutes,
                "Fan": a.subject or "",
                "Para": a.lesson_number or "",
                "Izoh": a.note or "",
            })
        return self._create_excel_file(pd.DataFrame(data), "Davomat")

    async def export_payments(self, group_id: Optional[int] = None) -> io.BytesIO:
        """Export payment report to Excel."""
        query = select(Student).options(joinedload(Student.group))
        if group_id:
            query = query.where(Student.group_id == group_id)
        query = query.where(Student.is_active == True).order_by(Student.name)

        result = await self.db.execute(query)
        students = result.unique().scalars().all()

        data = []
        for s in students:
            data.append({
                "ID": s.student_id,
                "F.I.O": s.name,
                "Guruh": s.group.name if s.group else "",
                "Kontrakt summasi": float(s.contract_amount),
                "To'langan": float(s.contract_paid),
                "Qolgan qarzi": float(s.contract_remaining),
                "To'lov %": f"{s.contract_percentage:.1f}%",
                "Holat": "To'liq to'langan" if s.is_contract_paid else "Qarzdor",
            })
        return self._create_excel_file(pd.DataFrame(data), "To'lovlar")

    async def export_groups(self) -> io.BytesIO:
        """Export groups to Excel."""
        query = select(Group).order_by(Group.name)
        result = await self.db.execute(query)
        groups = result.scalars().all()

        count_query = select(
            Student.group_id,
            func.count(Student.id).label("cnt")
        ).group_by(Student.group_id)
        count_result = await self.db.execute(count_query)
        counts = {row.group_id: row.cnt for row in count_result}

        data = []
        for g in groups:
            data.append({
                "ID": g.id,
                "Nomi": g.name,
                "Fakultet": g.faculty or "",
                "Kurs": g.course_year or "",
                "Talabalar soni": counts.get(g.id, 0),
                "Faol": "Ha" if g.is_active else "Yo'q",
            })
        return self._create_excel_file(pd.DataFrame(data), "Guruhlar ro'yxati")

    async def export_schedules(self, group_id: int) -> io.BytesIO:
        """Export schedule for a group to Excel."""
        query = (
            select(Schedule)
            .where(Schedule.group_id == group_id)
            .order_by(Schedule.day_of_week, Schedule.lesson_number)
        )
        result = await self.db.execute(query)
        schedules = result.scalars().all()

        day_names = {
            "monday": "Dushanba", "tuesday": "Seshanba", "wednesday": "Chorshanba",
            "thursday": "Payshanba", "friday": "Juma", "saturday": "Shanba",
            "sunday": "Yakshanba",
        }
        data = []
        for s in schedules:
            dv = s.day_of_week.value if s.day_of_week else ""
            data.append({
                "Kun": day_names.get(dv, dv),
                "Para": s.lesson_number or "",
                "Fan": s.subject or "",
                "O'qituvchi": s.teacher_name or "",
                "Xona": s.room or "",
                "Vaqt": (
                    f"{s.start_time.strftime('%H:%M') if s.start_time else ''}"
                    f" - "
                    f"{s.end_time.strftime('%H:%M') if s.end_time else ''}"
                ),
            })
        return self._create_excel_file(pd.DataFrame(data), "Dars jadvali")

    async def export_report(self, report_id: int) -> io.BytesIO:
        """Export a single report data to Excel."""
        from app.models.report import Report
        result = await self.db.execute(select(Report).where(Report.id == report_id))
        report = result.scalar_one_or_none()
        if not report:
            raise BadRequestException("Hisobot topilmadi")

        data = [{
            "ID": report.id,
            "Turi": report.report_type.value if report.report_type else "",
            "Nomi": report.name or "",
            "Holati": report.status.value if report.status else "",
            "Yaratilgan": report.created_at.strftime("%d.%m.%Y %H:%M") if report.created_at else "",
        }]
        return self._create_excel_file(pd.DataFrame(data), f"Hisobot #{report_id}")

    # ══════════════════════════════════════════════════════
    # STYLED EXCEL FILE CREATOR
    # ══════════════════════════════════════════════════════

    def _create_excel_file(self, df: pd.DataFrame, title: str) -> io.BytesIO:
        """Create styled Excel file from DataFrame."""
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ncols = max(len(df.columns), 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(2, 1, f"Sana: {now_tashkent().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(2, 1).alignment = Alignment(horizontal="right")

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(4, col_idx, col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(df.values, 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = thin_border
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'

        for col_idx, col_name in enumerate(df.columns, 1):
            max_length = len(str(col_name))
            for val in df[col_name]:
                max_length = max(max_length, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════
    # TEMPLATES
    # ══════════════════════════════════════════════════════

    async def get_import_template(self, template_type: str) -> io.BytesIO:
        """Get import template Excel file."""
        templates = {
            "students": {
                "columns": [
                    "F.I.O", "Telefon", "Email", "Tug'ilgan sana", "Jinsi",
                    "Manzil", "Pasport", "JSHSHIR", "Guruh", "Kontrakt", "To'langan",
                ],
                "sample": [{
                    "F.I.O": "Namuna Talaba", "Telefon": "+998901234567",
                    "Email": "talaba@email.com", "Tug'ilgan sana": "01.01.2000",
                    "Jinsi": "Erkak", "Manzil": "Toshkent sh.", "Pasport": "AA1234567",
                    "JSHSHIR": "12345678901234", "Guruh": "101-guruh",
                    "Kontrakt": 15000000, "To'langan": 5000000,
                }],
            },
            "attendance": {
                "columns": ["Talaba ID", "Sana", "Holat", "Kechikish", "Fan", "Izoh"],
                "sample": [{
                    "Talaba ID": "ST-2024-0001",
                    "Sana": now_tashkent().strftime("%d.%m.%Y"),
                    "Holat": "Keldi", "Kechikish": 0, "Fan": "Matematika", "Izoh": "",
                }],
            },
            "groups": {
                "columns": ["Nomi", "Fakultet", "Kurs"],
                "sample": [{"Nomi": "KI-25-01", "Fakultet": "Informatika", "Kurs": 1}],
            },
            "schedules": {
                "columns": [
                    "Guruh", "Kun", "Para", "Fan", "O'qituvchi",
                    "Xona", "Boshlanish", "Tugash",
                ],
                "sample": [{
                    "Guruh": "KI-25-01", "Kun": "Dushanba", "Para": 1,
                    "Fan": "Matematika", "O'qituvchi": "A.Aliyev", "Xona": "301",
                    "Boshlanish": "08:30", "Tugash": "10:00",
                }],
            },
        }

        t = templates.get(template_type)
        if not t:
            raise BadRequestException(f"Noma'lum shablon turi: {template_type}")

        df = pd.DataFrame(t["sample"], columns=t["columns"])
        return self._create_excel_file(df, f"{template_type}_template")

    # ══════════════════════════════════════════════════════
    # IMPORT: KONTINGENT (Students + Users + Groups) — UPSERT
    # ══════════════════════════════════════════════════════

    async def import_kontingent(
        self,
        file_data: bytes,
        update_existing: bool = True,
        create_users: bool = True,
        default_password: str = "12345678",
        deactivate_missing: bool = False,
    ) -> Dict[str, Any]:
        """
        Import students from Kontingent Excel file with FULL UPSERT logic.

        - New students -> INSERT (+ create User account)
        - Existing students (by student_id) -> UPDATE name, group, phone, passport, etc.
        - Students in DB but NOT in file -> optionally deactivate
        - New groups -> auto-create
        - Group changes -> update student's group_id

        Handles year-end scenarios: students move to new groups,
        new kontingent file has updated group assignments.
        """
        from app.models.user import User, UserRole
        from app.core.security import get_password_hash

        wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        # Pre-hash password once
        hashed_password = get_password_hash(default_password)

        # Load existing data in bulk
        existing_students_result = await self.db.execute(
            select(Student.id, Student.student_id, Student.name, Student.group_id, Student.is_active)
        )
        existing_students = {row.student_id: row for row in existing_students_result.fetchall()}
        all_student_ids_in_db = set(existing_students.keys())

        existing_users_result = await self.db.execute(select(User.login))
        existing_user_logins = set(row[0] for row in existing_users_result.fetchall())

        groups_result = await self.db.execute(select(Group))
        db_groups = groups_result.scalars().all()
        groups_cache = {g.name: g.id for g in db_groups}
        groups_lookup = build_group_lookup(db_groups)

        # Parse all rows (skip header rows — row 1 & 2)
        rows = list(ws.iter_rows(min_row=3, values_only=True))

        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        groups_created = 0
        users_created = 0
        errors = []
        student_ids_in_file: Set[str] = set()

        # Batch collections
        new_groups: Dict[str, Dict[str, Any]] = {}
        student_upserts: List[Dict[str, Any]] = []
        new_users: List[Dict[str, Any]] = []

        for row_idx, row in enumerate(rows, start=3):
            try:
                if not row or len(row) < 2:
                    continue

                student_id = str(row[0] or "").strip()
                full_name = str(row[1] or "").strip()
                if not student_id or not full_name:
                    continue

                student_ids_in_file.add(student_id)

                # Parse fields from kontingent columns
                passport = str(row[3] or "").strip() if len(row) > 3 else ""
                jshshir = str(row[4] or "").strip() if len(row) > 4 else ""
                birth_date_raw = row[6] if len(row) > 6 else None
                phone = str(row[7] or "").strip() if len(row) > 7 else ""
                specialty_name = str(row[12] or "").strip() if len(row) > 12 else ""
                course = row[13] if len(row) > 13 else None
                group_name = str(row[14] or "").strip() if len(row) > 14 else ""

                # Address fields
                country = str(row[15] or "").strip() if len(row) > 15 else ""
                region = str(row[16] or "").strip() if len(row) > 16 else ""
                district = str(row[17] or "").strip() if len(row) > 17 else ""
                address_detail = str(row[18] or "").strip() if len(row) > 18 else ""
                full_living_address = str(row[22] or "").strip() if len(row) > 22 else ""
                living_address = str(row[21] or "").strip() if len(row) > 21 else ""
                commute = str(row[23] or "").strip() if len(row) > 23 else ""

                address_parts = [p for p in [country, region, district, address_detail] if p]
                full_address = ", ".join(address_parts) if address_parts else full_living_address or living_address

                # Parse birth_date
                parsed_birth_date = None
                if birth_date_raw:
                    if isinstance(birth_date_raw, datetime):
                        parsed_birth_date = birth_date_raw.date()
                    elif isinstance(birth_date_raw, date):
                        parsed_birth_date = birth_date_raw
                    elif isinstance(birth_date_raw, str):
                        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                            try:
                                parsed_birth_date = datetime.strptime(birth_date_raw.strip(), fmt).date()
                                break
                            except Exception:
                                pass

                # Parse course number
                course_num = 1
                if course:
                    if isinstance(course, (int, float)):
                        course_num = int(course)
                    elif isinstance(course, str):
                        match = re.search(r'(\d+)', str(course))
                        if match:
                            course_num = int(match.group(1))

                # Resolve group — fuzzy match against existing DB groups
                group_id = None
                if group_name:
                    # Exact cache hit
                    if group_name in groups_cache:
                        group_id = groups_cache[group_name]
                    else:
                        # Fuzzy match
                        matched = fuzzy_match_group(group_name, {g.name: g for g in db_groups})
                        if matched:
                            group_id = groups_cache[matched]
                        elif group_name not in new_groups:
                            # Mark for creation
                            new_groups[group_name] = {
                                "name": group_name,
                                "faculty": specialty_name or "Noma'lum",
                                "course_year": course_num,
                                "is_active": True,
                            }
                            groups_created += 1

                # Determine UPSERT or INSERT
                is_existing = student_id in existing_students

                student_upserts.append({
                    "student_id": student_id,
                    "name": full_name,
                    "phone": phone or None,
                    "passport": passport or None,
                    "jshshir": jshshir or None,
                    "birth_date": parsed_birth_date,
                    "address": full_address or None,
                    "commute": commute or None,
                    "group_name": group_name,
                    "group_id": group_id,
                    "is_active": True,
                    "is_existing": is_existing,
                })

                if is_existing:
                    updated += 1
                else:
                    imported += 1
                    # Create user if needed
                    if create_users and student_id not in existing_user_logins:
                        new_users.append({
                            "login": student_id,
                            "password_hash": hashed_password,
                            "name": full_name,
                            "phone": phone or None,
                            "role": UserRole.STUDENT,
                            "is_active": True,
                            "is_first_login": True,
                        })
                        existing_user_logins.add(student_id)
                        users_created += 1

            except Exception as e:
                failed += 1
                errors.append({
                    "row": row_idx,
                    "student_id": str(row[0] if row and len(row) > 0 else ""),
                    "error": str(e)
                })

        wb.close()

        # ── BATCH WRITE ──

        # 1. Create new groups
        if new_groups:
            for gdata in new_groups.values():
                g = Group(**gdata)
                self.db.add(g)
            await self.db.flush()
            # Refresh cache
            new_groups_result = await self.db.execute(select(Group.id, Group.name))
            groups_cache = {row[1]: row[0] for row in new_groups_result.fetchall()}

        # 2. Upsert students
        for sdata in student_upserts:
            # Resolve group_id for newly created groups
            if sdata["group_id"] is None and sdata["group_name"]:
                sdata["group_id"] = groups_cache.get(sdata["group_name"])

            if sdata["is_existing"]:
                # UPDATE existing student
                update_fields: Dict[str, Any] = {
                    "name": sdata["name"],
                    "is_active": True,
                }
                if sdata["group_id"]:
                    update_fields["group_id"] = sdata["group_id"]
                if sdata["phone"]:
                    update_fields["phone"] = sdata["phone"]
                if sdata["passport"]:
                    update_fields["passport"] = sdata["passport"]
                if sdata["jshshir"]:
                    update_fields["jshshir"] = sdata["jshshir"]
                if sdata["birth_date"]:
                    update_fields["birth_date"] = sdata["birth_date"]
                if sdata["address"]:
                    update_fields["address"] = sdata["address"]
                if sdata["commute"]:
                    update_fields["commute"] = sdata["commute"]

                set_clause = ", ".join(f"{k} = :{k}" for k in update_fields)
                update_fields["sid"] = sdata["student_id"]
                await self.db.execute(
                    text(f"UPDATE students SET {set_clause} WHERE student_id = :sid"),
                    update_fields
                )
            else:
                # INSERT new student
                student = Student(
                    student_id=sdata["student_id"],
                    name=sdata["name"],
                    phone=sdata["phone"],
                    passport=sdata["passport"],
                    jshshir=sdata["jshshir"],
                    birth_date=sdata["birth_date"],
                    address=sdata["address"],
                    commute=sdata["commute"],
                    group_id=sdata["group_id"],
                    is_active=True,
                    contract_amount=0,
                    contract_paid=0,
                )
                self.db.add(student)

        # 3. Create users
        from app.models.user import User as UserModel, UserRole as UR
        for udata in new_users:
            user = UserModel(**udata)
            self.db.add(user)

        # 4. Deactivate missing students (if requested)
        deactivated = 0
        if deactivate_missing and student_ids_in_file:
            missing_ids = all_student_ids_in_db - student_ids_in_file
            if missing_ids:
                await self.db.execute(
                    text("UPDATE students SET is_active = false WHERE student_id = ANY(:ids)"),
                    {"ids": list(missing_ids)}
                )
                deactivated = len(missing_ids)

        await self.db.commit()

        # 5. Link users to students
        if new_users:
            new_logins = [u["login"] for u in new_users]
            users_result = await self.db.execute(
                select(UserModel.id, UserModel.login).where(UserModel.login.in_(new_logins))
            )
            user_id_map = {row[1]: row[0] for row in users_result.fetchall()}

            students_result = await self.db.execute(
                select(Student.id, Student.student_id).where(Student.student_id.in_(new_logins))
            )
            for student_db_id, student_sid in students_result.fetchall():
                if student_sid in user_id_map:
                    await self.db.execute(
                        text("UPDATE students SET user_id = :uid WHERE id = :sid"),
                        {"uid": user_id_map[student_sid], "sid": student_db_id}
                    )
            await self.db.commit()

        return {
            "success": True,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "failed": failed,
            "groups_created": groups_created,
            "users_created": users_created,
            "deactivated": deactivated,
            "errors": errors[:50],
            "message": (
                f"Kontingent: {imported} ta yangi, {updated} ta yangilandi, "
                f"{groups_created} ta guruh yaratildi, {users_created} ta foydalanuvchi yaratildi"
                + (f", {deactivated} ta nofaol qilindi" if deactivated else "")
            ),
        }

    # ══════════════════════════════════════════════════════
    # IMPORT: SCHEDULE from Excel — UPSERT with fuzzy group matching
    # ══════════════════════════════════════════════════════

    async def import_schedules(
        self,
        file_data: bytes,
        academic_year: str = "2025-2026",
        semester: int = 2,
        clear_existing: bool = False,
        use_ai: bool = True,
    ) -> Dict[str, Any]:
        """
        Import schedules from Excel with true UPSERT logic + AI enhancement.

        Supports two Excel formats:

        FORMAT 1 - Flat table:
        | Guruh | Kun | Para | Fan | O'qituvchi | Xona | Boshlanish | Tugash |

        FORMAT 2 - Grid (Google Sheets style):
        Groups as column headers, days as rows, lessons in cells.

        Features:
        - Fuzzy group name matching (KI-25-09 <-> KI_25-09)
        - AI-powered group matching for unresolved groups
        - AI-powered cell parsing for messy content
        - AI quality analysis with suggestions
        - TRUE UPSERT: updates existing schedules, adds new ones (never deletes)
        - Match key: (group_id, day_of_week, lesson_number, semester, academic_year)
        - Reports unmatched groups so admin can fix
        - Auto-detects format
        """
        try:
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise BadRequestException(f"Excel faylni o'qib bo'lmadi: {str(e)}")

        # Load groups from DB
        groups_result = await self.db.execute(select(Group))
        db_groups = groups_result.scalars().all()
        group_name_to_id = {g.name: g.id for g in db_groups}
        group_lookup = {g.name: g for g in db_groups}
        db_group_names = [g.name for g in db_groups]

        all_records: List[Dict[str, Any]] = []
        matched_groups: Set[str] = set()
        unmatched_groups: Set[str] = set()
        group_name_map: Dict[str, str] = {}  # sheet_name -> db_name

        # Sheets to skip (not schedule data)
        skip_sheet_keywords = ["bandligi", "bandlik", "o'qituvchi", "oqituvchi", "teacher"]

        for ws in wb.worksheets:
            # Skip non-schedule sheets
            sheet_lower = ws.title.lower().strip()
            if any(kw in sheet_lower for kw in skip_sheet_keywords):
                logger.info(f"Skipping non-schedule sheet: '{ws.title}'")
                continue

            logger.info(f"Parsing sheet: '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
            records = self._parse_schedule_sheet(ws, group_lookup, group_name_map)
            logger.info(f"Sheet '{ws.title}': {len(records)} records parsed")
            # Log first 3 records for debug
            for r in records[:3]:
                logger.info(f"  Sample: group={r.get('sheet_group_name')}, day={r.get('day')}, para={r.get('lesson_number')}, subj={r.get('subject')}, teacher={r.get('teacher')}")
            all_records.extend(records)

        wb.close()

        # ── AI ENHANCEMENT ──
        ai_result = None
        if use_ai:
            try:
                from app.services.ai_schedule_agent import AIScheduleAgent
                agent = AIScheduleAgent()
                if agent.is_available():
                    logger.info("AI Schedule Agent activated for import enhancement")
                    ai_result = await agent.enhance_import(
                        records=all_records,
                        db_group_names=db_group_names,
                        group_lookup=group_lookup,
                        group_name_map=group_name_map,
                    )
                    all_records = ai_result["records"]
                else:
                    logger.info("AI not available (no API key), using regex-only mode")
            except Exception as e:
                logger.warning(f"AI enhancement failed, continuing without AI: {e}")

        # Separate matched vs unmatched
        for rec in all_records:
            if rec.get("group_id"):
                matched_groups.add(rec["db_group_name"])
            else:
                unmatched_groups.add(rec["sheet_group_name"])

        # ── TRUE UPSERT: Update existing + Insert new (never delete) ──
        updated = 0
        inserted = 0
        skipped = 0
        errors: List[str] = []

        for rec in all_records:
            if not rec.get("group_id") or not rec.get("day") or not rec.get("subject"):
                skipped += 1
                continue

            try:
                start_time = rec.get("start_time")
                end_time = rec.get("end_time")
                lesson_num = rec.get("lesson_number", 1)

                # Use default time slots if not provided
                if not start_time or not end_time:
                    slot = DEFAULT_TIME_SLOTS.get(lesson_num, DEFAULT_TIME_SLOTS[1])
                    start_time = start_time or parse_time(slot[0])
                    end_time = end_time or parse_time(slot[1])

                # Check if schedule already exists by unique key
                existing_result = await self.db.execute(
                    select(Schedule).where(
                        Schedule.group_id == rec["group_id"],
                        Schedule.day_of_week == rec["day"],
                        Schedule.lesson_number == lesson_num,
                        Schedule.semester == semester,
                        Schedule.academic_year == academic_year,
                    )
                )
                existing = existing_result.scalars().first()

                if existing:
                    # UPDATE existing schedule
                    existing.subject = rec["subject"]
                    existing.schedule_type = rec.get("schedule_type", ScheduleType.LECTURE)
                    existing.start_time = start_time
                    existing.end_time = end_time
                    existing.room = rec.get("room")
                    existing.building = rec.get("building")
                    existing.teacher_name = rec.get("teacher")
                    existing.is_active = True
                    existing.is_cancelled = False
                    updated += 1
                else:
                    # INSERT new schedule
                    schedule = Schedule(
                        group_id=rec["group_id"],
                        subject=rec["subject"],
                        schedule_type=rec.get("schedule_type", ScheduleType.LECTURE),
                        day_of_week=rec["day"],
                        start_time=start_time,
                        end_time=end_time,
                        lesson_number=lesson_num,
                        room=rec.get("room"),
                        building=rec.get("building"),
                        teacher_name=rec.get("teacher"),
                        semester=semester,
                        academic_year=academic_year,
                        is_active=True,
                    )
                    self.db.add(schedule)
                    inserted += 1

            except Exception as e:
                skipped += 1
                errors.append(f"{rec.get('sheet_group_name', '?')}: {str(e)}")

        await self.db.commit()

        synced = updated + inserted

        # Build response
        response = {
            "success": True,
            "total_records": len(all_records),
            "synced": synced,
            "updated": updated,
            "inserted": inserted,
            "skipped": skipped,
            "matched_groups": list(matched_groups),
            "unmatched_groups": list(unmatched_groups),
            "group_name_map": group_name_map,
            "errors": errors[:20],
            "message": (
                f"Jadval: {updated} ta yangilandi, {inserted} ta qo'shildi, {skipped} ta o'tkazib yuborildi. "
                f"{len(matched_groups)} ta guruh topildi, {len(unmatched_groups)} ta topilmadi."
            ),
        }

        # Add AI results to response
        if ai_result:
            ai_matched = ai_result.get("ai_matched_groups", {})
            ai_parsed = ai_result.get("ai_parsed_cells", 0)
            analysis = ai_result.get("analysis")
            tokens = ai_result.get("tokens_used", 0)

            response["ai"] = {
                "enabled": True,
                "matched_groups": ai_matched,
                "matched_groups_count": len(ai_matched),
                "parsed_cells_count": ai_parsed,
                "tokens_used": tokens,
                "analysis": analysis,
            }

            # Update message with AI info
            if ai_matched or ai_parsed:
                ai_msg_parts = []
                if ai_matched:
                    ai_msg_parts.append(f"AI {len(ai_matched)} ta guruhni moslashtirdi")
                if ai_parsed:
                    ai_msg_parts.append(f"{ai_parsed} ta katakchani tahlil qildi")
                response["message"] += " | AI: " + ", ".join(ai_msg_parts) + "."
        else:
            response["ai"] = {"enabled": False}

        return response

    def _parse_schedule_sheet(
        self,
        ws,
        group_lookup: Dict[str, Any],
        group_name_map: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """
        Parse a single worksheet for schedule data.
        Auto-detects format:
        - If row 1 or row 2 has flat table headers ('Guruh', 'Fan', 'Kun') -> flat table
        - Otherwise -> Google Sheets grid format (groups as columns)
        """
        records: List[Dict[str, Any]] = []

        # Read first 2 rows to detect format
        flat_keywords = ["guruh", "group", "fan", "subject", "kun", "day"]
        
        for check_row in [1, 2]:
            header_vals = []
            for cell in ws[check_row]:
                header_vals.append(str(cell.value or "").strip().lower())
            joined = " ".join(header_vals)
            if any(kw in joined for kw in flat_keywords):
                records = self._parse_flat_schedule(ws, group_lookup, group_name_map, header_row=check_row)
                return records

        # Grid format
        records = self._parse_grid_schedule(ws, group_lookup, group_name_map)
        return records

    def _parse_flat_schedule(
        self,
        ws,
        group_lookup: Dict[str, Any],
        group_name_map: Dict[str, str],
        header_row: int = 1,
    ) -> List[Dict[str, Any]]:
        """
        Parse flat table format:
        | Guruh | Kun | Para | Fan | O'qituvchi | Xona | Bino | Boshlanish | Tugash |
        """
        records: List[Dict[str, Any]] = []

        # Map header columns
        headers: Dict[str, int] = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            val = str(cell.value or "").strip().lower()
            if val in ("guruh", "group", "guruh nomi"):
                headers["group"] = col_idx
            elif val in ("kun", "day", "hafta kuni"):
                headers["day"] = col_idx
            elif val in ("para", "lesson", "dars"):
                headers["lesson"] = col_idx
            elif val in ("fan", "subject", "fan nomi"):
                headers["subject"] = col_idx
            elif val in ("o'qituvchi", "teacher", "ustoz", "oqituvchi"):
                headers["teacher"] = col_idx
            elif val in ("xona", "room", "auditoriya"):
                headers["room"] = col_idx
            elif val in ("bino", "building"):
                headers["building"] = col_idx
            elif val in ("boshlanish", "start", "boshi"):
                headers["start"] = col_idx
            elif val in ("tugash", "end", "oxiri"):
                headers["end"] = col_idx
            elif val in ("tur", "type"):
                headers["type"] = col_idx

        if "group" not in headers or "subject" not in headers:
            return records

        type_map = {
            "ma'ruza": ScheduleType.LECTURE, "lecture": ScheduleType.LECTURE, "maruza": ScheduleType.LECTURE,
            "amaliy": ScheduleType.PRACTICE, "practice": ScheduleType.PRACTICE,
            "laboratoriya": ScheduleType.LAB, "lab": ScheduleType.LAB,
            "seminar": ScheduleType.SEMINAR,
            "imtihon": ScheduleType.EXAM, "exam": ScheduleType.EXAM,
            "konsultatsiya": ScheduleType.CONSULTATION, "consultation": ScheduleType.CONSULTATION,
        }

        data_start = header_row + 1
        for row_idx in range(data_start, ws.max_row + 1):
            group_val = ws.cell(row_idx, headers["group"]).value
            if not group_val:
                continue

            group_name = str(group_val).strip()
            subject_val = ws.cell(row_idx, headers.get("subject", 0)).value if "subject" in headers else None
            subject = str(subject_val or "").strip()
            if not subject:
                continue

            # Resolve group
            group_id = None
            db_group_name = None
            matched = fuzzy_match_group(group_name, group_lookup)
            if matched:
                group_id = group_lookup[matched].id
                db_group_name = matched
                group_name_map[group_name] = matched

            day = parse_day(ws.cell(row_idx, headers["day"]).value) if "day" in headers else None
            lesson_num = None
            if "lesson" in headers:
                lv = ws.cell(row_idx, headers["lesson"]).value
                if lv is not None:
                    try:
                        lesson_num = int(float(str(lv)))
                    except (ValueError, TypeError):
                        pass

            start_time = parse_time(ws.cell(row_idx, headers["start"]).value) if "start" in headers else None
            end_time = parse_time(ws.cell(row_idx, headers["end"]).value) if "end" in headers else None

            teacher = str(ws.cell(row_idx, headers["teacher"]).value or "").strip() if "teacher" in headers else None
            room = str(ws.cell(row_idx, headers["room"]).value or "").strip() if "room" in headers else None
            building = str(ws.cell(row_idx, headers["building"]).value or "").strip() if "building" in headers else None

            stype = ScheduleType.LECTURE
            if "type" in headers:
                tval = str(ws.cell(row_idx, headers["type"]).value or "").strip().lower()
                stype = type_map.get(tval, ScheduleType.LECTURE)

            records.append({
                "sheet_group_name": group_name,
                "db_group_name": db_group_name,
                "group_id": group_id,
                "day": day,
                "lesson_number": lesson_num or 1,
                "subject": subject,
                "teacher": teacher,
                "room": room,
                "building": building,
                "start_time": start_time,
                "end_time": end_time,
                "schedule_type": stype,
            })

        return records

    # Regex to parse cell content: "Subject (type) Teacher Room"
    _CELL_TYPE_RE = re.compile(
        r"^(.+?)\s*\((ma.ruza|amaliy|lo?baratoriya|seminar|imtihon|konsultatsiya|exam|lab)\)\s*(.*)",
        re.IGNORECASE,
    )
    # Regex to extract room+building from the rest after teacher name
    # Matches: "307-xona A bino", "L-2 B bino", "Sport zal A bino", "Faollar zali A bino"
    # Also matches: "101-xona", "Aud. 305", "xona 101", "507A", standalone room numbers
    _CELL_ROOM_RE = re.compile(
        r"^(.+?)\s+(\d{2,4}[A-Za-z]?(?:-xona)?(?:\s+[A-Z]\s+bino)?|"
        r"Sport\s+zal(?:\s+[A-Z]\s+bino)?|"
        r"L-\d+(?:\s+[A-Z]\s+bino)?|"
        r"Faollar\s+zali(?:\s+[A-Z]\s+bino)?|"
        r"Aud\.?\s*\d+|"
        r"xona[\s-]?\d+)$",
        re.IGNORECASE,
    )
    # Helper to split room and building from combined string
    _ROOM_BUILDING_RE = re.compile(
        r"^(.+?)\s+([A-Z])\s+bino$",
        re.IGNORECASE,
    )

    # ── NEW: Pattern for plain-text medical schedule cells ──
    # Matches: "... X bino NNN" or "... X bino" (room optional) at the end
    # Examples: "Anatomiya A bino 305", "Tib kimyo D bino 309", "TKK C bino"
    # Also handles no space before letter: "Lotin tiliA bino 504"
    _PLAIN_BINO_RE = re.compile(
        r'^(.+?)\s*([A-ZА-Я])\s+bino\s+(.+)$',
        re.IGNORECASE,
    )
    # Pattern for "X bino" WITHOUT room number (e.g., "TKK C bino")
    _PLAIN_BINO_NO_ROOM_RE = re.compile(
        r'^(.+?)\s+([A-ZА-Я])\s+bino\s*$',
        re.IGNORECASE,
    )

    # Known medical/university subject keywords for smarter splitting
    _KNOWN_SUBJECTS = [
        "anatomiya", "biologiya", "tib kimyo", "tibbiy kimyo", "kimyo",
        "fiziologiya", "normal fiziologiya", "patologik fiziologiya",
        "gistologiya", "farmakologiya", "mikrobiologiya", "biokimyo",
        "lotin tili", "ingliz tili", "rus tili", "nemis tili",
        "tkk", "pkrs", "odk", "dinshunoslik", "falsafa",
        "stomatologiya", "stomatologiyaga kirish",
        "farmatsiya", "farmatsevtika", "farmatsevtikada axborot texnologiyalari",
        "gigiyena", "jamoat salomatligi", "davolash ishi",
        "pediatriya", "akusherlik", "jarrohlik",
        "psixologiya", "pedagogika", "sotsiologiya",
        "informatika", "matematika", "fizika",
        "tarix", "iqtisodiyot", "huquqshunoslik",
        "noorganik kimyo", "organik kimyo", "analitik kimyo",
        "tibbiyotda xorijiy til", "tibbiyotda axborot texnologiyalari",
        "oʻzbekistonning eng yangi tarixi", "o'zbekiston tarixi",
        "direktor bilan uchrashuv",
        "valeologiya", "ekologiya", "genetika",
        "patologik anatomiya", "yuqumli kasalliklar",
        "umumiy jarrohlik", "ichki kasalliklar",
        "teri-tanosil kasalliklari", "nerv kasalliklari",
        "otorinolaringologiya", "oftalmologiya",
        "tibbiy fizika", "tibbiy biologiya",
        "milliy istiqlol g'oyasi", "milliy g'oya",
        "harbiy tayyorgarlik", "jismoniy tarbiya",
        "ommaviy axborot vositalari", "sotsiologiya",
    ]

    # Map type keywords to ScheduleType
    _CELL_TYPE_MAP = {
        "ma'ruza": ScheduleType.LECTURE,
        "ma`ruza": ScheduleType.LECTURE,
        "maruza": ScheduleType.LECTURE,
        "ma\u2018ruza": ScheduleType.LECTURE,
        "ma\u2019ruza": ScheduleType.LECTURE,
        "amaliy": ScheduleType.PRACTICE,
        "laboratoriya": ScheduleType.LAB,
        "lobaratoriya": ScheduleType.LAB,
        "lab": ScheduleType.LAB,
        "seminar": ScheduleType.SEMINAR,
        "imtihon": ScheduleType.EXAM,
        "exam": ScheduleType.EXAM,
        "konsultatsiya": ScheduleType.CONSULTATION,
        "qo'shimcha dars": ScheduleType.PRACTICE,
        "qo\u2018shimcha dars": ScheduleType.PRACTICE,
        "qo`shimcha dars": ScheduleType.PRACTICE,
    }

    # Type keywords for inline detection
    _TYPE_KEYWORDS_RE = re.compile(
        r"\b(ma['\`\u2018\u2019]?ruza|amaliy|lo?baratoriya|lab|seminar|imtihon|exam|konsultatsiya|qo['\`\u2018\u2019]?shimcha\s+dars)\b",
        re.IGNORECASE,
    )

    def _split_room_building(self, raw_room: str):
        """Split '307-xona A bino' into room='307-xona' and building='A bino'."""
        if not raw_room:
            return raw_room, None
        m = self._ROOM_BUILDING_RE.match(raw_room.strip())
        if m:
            return m.group(1).strip(), f"{m.group(2)} bino"
        return raw_room.strip(), None

    def _find_known_subject(self, text: str):
        """
        Try to find a known subject at the start of the text.
        Returns (subject, rest_of_text) or (None, text).
        """
        text_lower = text.lower().strip()
        # Sort by length descending to match longest first
        for subj in sorted(self._KNOWN_SUBJECTS, key=len, reverse=True):
            if text_lower.startswith(subj):
                rest = text[len(subj):].strip()
                subject = text[:len(subj)].strip()
                return subject, rest
        return None, text

    def _split_room_and_teacher_after_bino(self, after_bino: str):
        """
        Split room and teacher from text after "X bino".

        Patterns:
        - "305" → room="305", teacher=None
        - "312 / 323" → room="312 / 323", teacher=None
        - "312 / 323 Abdurahmonov Sh / Abduraxmonov T" → room="312 / 323", teacher="Abdurahmonov Sh / Abduraxmonov T"
        - "511/401 To'ychiyeva Sh Turg'unova D" → room="511/401", teacher="To'ychiyeva Sh / Turg'unova D"
        - "Faollar zali" → room="Faollar zali", teacher=None
        - "06 B" → room="06 B", teacher=None
        - "A guruhga A bino 305" → room="A guruhga A bino 305", teacher=None (edge case)

        Returns (room, teacher)
        """
        if not after_bino:
            return None, None

        text = after_bino.strip().rstrip('/')  # Remove trailing slash

        # Special named rooms
        named_rooms = ['faollar zali', 'sport zal', 'a guruhga', 'b guruhga']
        for nr in named_rooms:
            if text.lower().startswith(nr):
                return text, None

        # Try to split: rooms are numbers/digits, teachers are names (start with uppercase letter)
        # Pattern: room part = digits/slashes at start, teacher part = names after
        # Examples: "312 / 323 Abdurahmonov Sh / Abduraxmonov T"
        #           "511/401 To'ychiyeva Sh"
        #           "305"
        #           "06 B"
        #           "404 / 406 Jo'raboyev A / Sidiqjanov N"

        # Strategy: scan tokens, room tokens are digits or single letters after digits or "/"
        # Once we hit a word that looks like a name (not a number, not "/" , not single letter), that starts the teacher
        tokens = text.split()
        room_parts = []
        teacher_start_idx = None

        i = 0
        while i < len(tokens):
            token = tokens[i]
            clean = token.strip('/')

            # Is this a room-like token?
            if re.match(r'^\d+[A-Za-z]?/?$', clean):
                room_parts.append(token)
                i += 1
                continue
            # Slash separator between rooms
            if token == '/':
                room_parts.append(token)
                i += 1
                continue
            # Single uppercase letter after room number (like "06 B")
            if len(clean) == 1 and clean.isupper() and room_parts:
                room_parts.append(token)
                i += 1
                continue
            # Room with slash like "511/401" or "602/603"
            if re.match(r'^\d+/\d+$', clean):
                room_parts.append(token)
                i += 1
                continue

            # This token doesn't look like a room → start of teacher
            teacher_start_idx = i
            break
            i += 1

        if room_parts:
            room = " ".join(room_parts).strip().rstrip('/')
        else:
            room = text  # Fallback: entire text is room

        if teacher_start_idx is not None and teacher_start_idx < len(tokens):
            teacher = " ".join(tokens[teacher_start_idx:]).strip().rstrip('/')
            if teacher and len(teacher) >= 2:
                return room, teacher

        return room, None

    def _parse_cell_content(self, raw: str):
        """
        Parse schedule cell content — supports ALL formats including plain text.

        Formats supported:
        1. "Subject (type) Teacher Room"       — parenthesized type
        2. "Subject\\nTeacher\\nRoom"           — newline-separated
        3. "Subject Teacher X bino Room"       — plain text with "X bino NNN" pattern
        4. "Subject type Teacher X bino Room"  — inline type keyword (ma'ruza, amaliy, lab)
        5. "Subject X bino Room"               — no teacher, just subject + location
        6. "Subject Teacher Room"              — old regex-based fallback
        7. "Subject"                           — subject only

        Returns (subject, schedule_type, teacher, room, building)
        """
        text = raw.strip()
        if not text or text in ("-", "—", " ", ","):
            return None, None, None, None, None

        # Normalize multiple spaces to single space
        text = re.sub(r'\s{2,}', ' ', text)

        # --- Format 2: newline-separated ---
        if "\n" in text:
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            subject = lines[0] if len(lines) > 0 else None
            teacher = lines[1] if len(lines) > 1 else None
            raw_room = lines[2] if len(lines) > 2 else None
            room, building = self._split_room_building(raw_room) if raw_room else (None, None)

            # Check if subject line itself contains type in parentheses
            stype = ScheduleType.LECTURE
            if subject:
                m = self._CELL_TYPE_RE.match(subject)
                if m:
                    subject = m.group(1).strip()
                    stype_raw = m.group(2).strip().lower().replace("\u2018", "'").replace("\u2019", "'")
                    stype = self._CELL_TYPE_MAP.get(stype_raw, ScheduleType.LECTURE)
                    extra = m.group(3).strip()
                    if extra and not teacher:
                        teacher = extra

            return subject, stype, teacher, room, building

        # --- Format 1: "Subject (type) Teacher Room" ---
        m = self._CELL_TYPE_RE.match(text)
        if m:
            subject = m.group(1).strip()
            stype_raw = m.group(2).strip().lower().replace("\u2018", "'").replace("\u2019", "'")
            stype = self._CELL_TYPE_MAP.get(stype_raw, ScheduleType.LECTURE)
            rest = m.group(3).strip()

            teacher = None
            room = None
            building = None
            if rest:
                # Try "X bino NNN" at end of rest
                bm = self._PLAIN_BINO_RE.match(rest)
                if bm:
                    teacher = bm.group(1).strip() or None
                    building = f"{bm.group(2).upper()} bino"
                    room = bm.group(3).strip()
                else:
                    rm = self._CELL_ROOM_RE.match(rest)
                    if rm:
                        teacher = rm.group(1).strip()
                        room, building = self._split_room_building(rm.group(2).strip())
                    else:
                        teacher = rest
            return subject, stype, teacher, room, building

        # --- Format 3/4/5: Plain text with "X bino NNN" pattern ---
        # This is the main pattern for medical schedule cells
        bm = self._PLAIN_BINO_RE.match(text)
        if bm:
            before_bino = bm.group(1).strip()  # Everything before "X bino"
            bino_letter = bm.group(2).upper()
            after_bino = bm.group(3).strip()    # Everything after "X bino"
            building = f"{bino_letter} bino"

            # Parse after_bino: can be "305", "312 / 323", "312 / 323 Teacher1 / Teacher2"
            # or "Faollar zali", "A guruhga A bino 305"
            room, after_teacher = self._split_room_and_teacher_after_bino(after_bino)

            # Now parse "before_bino" to extract subject, optional type, optional teacher
            subject = None
            before_teacher = None
            stype = ScheduleType.LECTURE

            # Check for inline type keyword
            type_match = self._TYPE_KEYWORDS_RE.search(before_bino)
            if type_match:
                subject = before_bino[:type_match.start()].strip()
                stype_raw = type_match.group(1).strip().lower().replace("\u2018", "'").replace("\u2019", "'")
                stype = self._CELL_TYPE_MAP.get(stype_raw, ScheduleType.LECTURE)
                before_teacher = before_bino[type_match.end():].strip() or None
            else:
                # Try known subjects
                known_subj, rest_after_subj = self._find_known_subject(before_bino)
                if known_subj:
                    subject = known_subj
                    before_teacher = rest_after_subj.strip() or None
                    if before_teacher and len(before_teacher.strip()) < 2:
                        before_teacher = None
                else:
                    # Heuristic splitting for unknown subjects
                    words = before_bino.split()
                    if len(words) <= 2:
                        subject = before_bino
                    else:
                        subject = before_bino
                        for split_at in range(len(words) - 1, 0, -1):
                            potential_teacher = " ".join(words[split_at:])
                            potential_subject = " ".join(words[:split_at])
                            first_char = potential_teacher[0] if potential_teacher else ''
                            if first_char.isupper() and len(potential_teacher) > 2:
                                ps_lower = potential_subject.lower().strip()
                                is_known = any(ps_lower == k or ps_lower.startswith(k) for k in self._KNOWN_SUBJECTS)
                                if is_known:
                                    subject = potential_subject
                                    before_teacher = potential_teacher
                                    break

            # Combine teacher from before_bino and after_bino
            teacher = None
            if before_teacher and after_teacher:
                teacher = f"{before_teacher} / {after_teacher}" if before_teacher != after_teacher else before_teacher
            elif before_teacher:
                teacher = before_teacher
            elif after_teacher:
                teacher = after_teacher

            return subject, stype, teacher, room, building

        # --- Format 3b: "Subject X bino" — no room number ---
        bm_no_room = self._PLAIN_BINO_NO_ROOM_RE.match(text)
        if bm_no_room:
            before_bino = bm_no_room.group(1).strip()
            bino_letter = bm_no_room.group(2).upper()
            building = f"{bino_letter} bino"
            # Parse before_bino same way
            type_match = self._TYPE_KEYWORDS_RE.search(before_bino)
            if type_match:
                subject = before_bino[:type_match.start()].strip()
                stype_raw = type_match.group(1).strip().lower().replace("\u2018", "'").replace("\u2019", "'")
                stype = self._CELL_TYPE_MAP.get(stype_raw, ScheduleType.LECTURE)
                teacher = before_bino[type_match.end():].strip() or None
                return subject, stype, teacher, None, building
            known_subj, rest = self._find_known_subject(before_bino)
            if known_subj:
                teacher = rest.strip() or None
                if teacher and len(teacher.strip()) < 2:
                    teacher = None
                return known_subj, ScheduleType.LECTURE, teacher, None, building
            return before_bino, ScheduleType.LECTURE, None, None, building

        # --- Format 6: Try old room pattern at end (without "bino") ---
        room_at_end = re.search(
            r'\s+(\d{2,4}[A-Za-z]?-xona(?:\s+[A-Z]\s+bino)?|'
            r'Sport\s+zal(?:\s+[A-Z]\s+bino)?|'
            r'L-\d+(?:\s+[A-Z]\s+bino)?|'
            r'Faollar\s+zali(?:\s+[A-Z]\s+bino)?)$',
            text, re.IGNORECASE
        )
        if room_at_end:
            subject = text[:room_at_end.start()].strip()
            room, building = self._split_room_building(room_at_end.group(1).strip())
            return subject, ScheduleType.LECTURE, None, room, building

        # --- Format 7: Subject only ---
        # Check for inline type keyword even without room
        type_match = self._TYPE_KEYWORDS_RE.search(text)
        if type_match:
            subject = text[:type_match.start()].strip()
            stype_raw = type_match.group(1).strip().lower().replace("\u2018", "'").replace("\u2019", "'")
            stype = self._CELL_TYPE_MAP.get(stype_raw, ScheduleType.LECTURE)
            teacher = text[type_match.end():].strip() or None
            return subject, stype, teacher, None, None

        return text, ScheduleType.LECTURE, None, None, None

    def _get_merged_cell_value(self, ws, row, col):
        """
        Get cell value considering merged cells.
        If a cell is part of a merged range, return the value from the top-left cell.
        """
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return ws.cell(row, col).value

    def _parse_grid_schedule(
        self,
        ws,
        group_lookup: Dict[str, Any],
        group_name_map: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """
        Parse grid format with FULL merged cell support.

        Excel structure (per block):
        - Row 1: department/direction name (merged) — SKIPPED
        - Row 2: group names as column headers
        - Column A (or block-local): Day of week (merged across lesson rows)
        - Column B (or block-local): Lesson number
        - Column C (or block-local): Time range
        - Data columns: Cell content = "Subject (type) Teacher Room Building"

        Supports:
        - Multiple blocks in one sheet (2-bosqich style: each faculty has own day/lesson/time cols)
        - Merged data cells spanning multiple group columns (same lesson for multiple groups)
        - Groups in rows 2, 3, 4, or 1
        - Merged cells for days spanning multiple lesson rows
        """
        records: List[Dict[str, Any]] = []

        # --- Build merged cell lookups ---
        # 1) Standard merged lookup for meta cells (day, lesson, time)
        merged_lookup = {}  # (row, col) -> value from top-left
        # 2) Data merge ranges for distributing one cell value to multiple group columns
        data_merge_ranges = []  # list of (min_row, max_row, min_col, max_col, value)

        for merged_range in ws.merged_cells.ranges:
            top_val = ws.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_lookup[(r, c)] = top_val
            # Save data merge ranges (we'll use these for data cells)
            data_merge_ranges.append((
                merged_range.min_row, merged_range.max_row,
                merged_range.min_col, merged_range.max_col,
                top_val
            ))

        def get_cell(row, col):
            """Get cell value, resolving merged cells."""
            if (row, col) in merged_lookup:
                return merged_lookup[(row, col)]
            return ws.cell(row, col).value

        # --- Detect group row: try rows 2, 3, 4, then row 1 ---
        group_cols: Dict[int, tuple] = {}  # col_idx -> (sheet_name, group_id, db_name)
        group_header_row = 2

        skip_values = {"kun", "day", "para", "vaqt", "time", "hafta kuni", "dars",
                        "soat", "№", "lesson", "hour", "", "dars vaqti"}

        # Stream/flow names to skip (not group names)
        stream_re = re.compile(r'^\d+-?\s*(oqim|поток|stream|bosqich|kurs)$', re.IGNORECASE)

        for try_row in [2, 3, 4]:
            found_groups = 0
            temp_cols = {}
            for col_idx in range(1, ws.max_column + 1):
                val = get_cell(try_row, col_idx)
                if not val:
                    continue
                name = str(val).strip()
                if name.lower() in skip_values or len(name) < 2:
                    continue
                # Skip stream/flow headers like "1-oqim", "2-oqim"
                if stream_re.match(name):
                    continue
                # Skip if same value as row 1 merged header (faculty name, not group)
                row1_val = get_cell(1, col_idx)
                if row1_val and str(row1_val).strip() == name:
                    continue
                # Check if it looks like a group name (has letters + digits)
                if re.search(r'[A-Za-zА-Яа-яЎўҚқҒғҲҳ]', name) and re.search(r'\d', name):
                    matched = fuzzy_match_group(name, group_lookup)
                    if matched:
                        temp_cols[col_idx] = (name, group_lookup[matched].id, matched)
                        group_name_map[name] = matched
                        found_groups += 1
                    else:
                        temp_cols[col_idx] = (name, None, None)
                        found_groups += 1

            if found_groups >= 1:
                group_cols = temp_cols
                group_header_row = try_row
                break

        # Fallback: try row 1
        if not group_cols:
            group_header_row = 1
            for col_idx in range(1, ws.max_column + 1):
                val = get_cell(1, col_idx)
                if not val:
                    continue
                name = str(val).strip()
                if name.lower() in skip_values or len(name) < 2:
                    continue
                # Skip stream/flow headers like "1-oqim", "2-oqim"
                if stream_re.match(name):
                    continue
                if re.search(r'[A-Za-zА-Яа-яЎўҚқҒғҲҳ]', name) and re.search(r'\d', name):
                    matched = fuzzy_match_group(name, group_lookup)
                    if matched:
                        group_cols[col_idx] = (name, group_lookup[matched].id, matched)
                        group_name_map[name] = matched
                    else:
                        group_cols[col_idx] = (name, None, None)

        if not group_cols:
            logger.warning(f"Grid parser: No groups found in sheet '{ws.title}'")
            return records

        data_start_row = group_header_row + 1

        logger.info(f"Grid parser '{ws.title}': group_header_row={group_header_row}, "
                     f"groups={len(group_cols)}, merged_cells={len(ws.merged_cells.ranges)}")

        # --- Detect BLOCKS: each block has its own day/lesson/time columns ---
        # A block = (day_col, lesson_col, time_col, set_of_group_col_indices)
        # In simple sheets: one block with cols A,B,C for all groups
        # In multi-faculty sheets: multiple blocks separated by "Dars vaqti" meta columns

        sorted_group_cols = sorted(group_cols.keys())

        # Find all meta column sets: columns in data rows that have day names or lesson numbers
        # A meta column is any column NOT in group_cols that has day/lesson/time data
        meta_col_sets = []  # list of (day_col, lesson_col, time_col)

        # Method: look for day names in data_start_row across all columns
        day_columns = []
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in group_cols:
                continue
            val = get_cell(data_start_row, col_idx)
            if val and parse_day(val):
                day_columns.append(col_idx)

        if not day_columns:
            # Fallback: just use col 1
            day_columns = [1]

        # For each day column, find the adjacent lesson and time columns
        for day_col in day_columns:
            lesson_col = None
            time_col = None
            # Check next 1-2 columns for lesson number and time
            for offset in [1, 2]:
                check_col = day_col + offset
                if check_col in group_cols:
                    break
                val = get_cell(data_start_row, check_col)
                if val is not None:
                    # Is it a number (lesson)?
                    try:
                        ln = int(float(str(val)))
                        if 1 <= ln <= 10:
                            lesson_col = check_col
                            continue
                    except (ValueError, TypeError):
                        pass
                    # Is it a time string?
                    val_str = str(val).strip()
                    if re.match(r'\d{1,2}:\d{2}', val_str):
                        time_col = check_col
            meta_col_sets.append((day_col, lesson_col, time_col))

        # Assign each group column to its nearest preceding meta block
        # Build blocks: each block = (day_col, lesson_col, time_col, [group_cols])
        blocks = []
        meta_col_sets_sorted = sorted(meta_col_sets, key=lambda x: x[0])

        for i, (day_c, lesson_c, time_c) in enumerate(meta_col_sets_sorted):
            # Find group columns that belong to this block:
            # All group columns after this meta and before the next meta
            if i + 1 < len(meta_col_sets_sorted):
                next_meta_start = meta_col_sets_sorted[i + 1][0]
            else:
                next_meta_start = ws.max_column + 1

            block_groups = {
                col: group_cols[col]
                for col in sorted_group_cols
                if col > (time_c or lesson_c or day_c) and col < next_meta_start
            }

            if block_groups:
                blocks.append((day_c, lesson_c, time_c, block_groups))

        # If no blocks were formed, create a single default block
        if not blocks:
            min_gc = min(sorted_group_cols)
            meta_cols = list(range(1, min_gc))
            day_c = meta_cols[0] if len(meta_cols) >= 1 else 1
            lesson_c = meta_cols[1] if len(meta_cols) >= 2 else None
            time_c = meta_cols[2] if len(meta_cols) >= 3 else None
            blocks = [(day_c, lesson_c, time_c, group_cols)]

        logger.info(f"Grid parser '{ws.title}': {len(blocks)} block(s) detected")
        for bi, (dc, lc, tc, bg) in enumerate(blocks):
            grp_names = [bg[c][0] for c in sorted(bg.keys())][:5]
            logger.info(f"  Block {bi}: day_col={dc}, lesson_col={lc}, time_col={tc}, "
                         f"groups={len(bg)} ({grp_names}{'...' if len(bg) > 5 else ''})")

        # --- Build reverse lookup: for each data cell merged range, which group columns it covers ---
        # This is needed because a merged data cell like D13:F13 means the lesson is for groups D, E, F
        # We index by (row, min_col) -> list of group cols covered
        merged_data_coverage = {}  # (min_row, min_col) -> set of group col indices covered
        group_col_set = set(group_cols.keys())
        for min_r, max_r, min_c, max_c, val in data_merge_ranges:
            if not val:
                continue
            covered_groups = group_col_set & set(range(min_c, max_c + 1))
            if covered_groups and len(covered_groups) > 1:
                for r in range(min_r, max_r + 1):
                    merged_data_coverage[(r, min_c)] = covered_groups

        # --- Parse data rows per block ---
        for day_col, lesson_col, time_col, block_groups in blocks:
            current_day = None
            current_lesson = None
            current_start_time = None
            current_end_time = None

            for row_idx in range(data_start_row, ws.max_row + 1):
                # Day of week (merged cells auto-resolved via get_cell)
                day_val = get_cell(row_idx, day_col)
                if day_val:
                    d = parse_day(day_val)
                    if d:
                        current_day = d

                # Lesson number
                if lesson_col:
                    lesson_val = get_cell(row_idx, lesson_col)
                    if lesson_val is not None:
                        try:
                            ln = int(float(str(lesson_val)))
                            if 1 <= ln <= 10:
                                current_lesson = ln
                        except (ValueError, TypeError):
                            pass

                # Time range
                if time_col:
                    time_val = get_cell(row_idx, time_col)
                    if time_val:
                        time_str = str(time_val).strip()
                        parts = re.split(r'[-–—]', time_str)
                        if len(parts) == 2:
                            st = parse_time(parts[0].strip())
                            et = parse_time(parts[1].strip())
                            if st:
                                current_start_time = st
                            if et:
                                current_end_time = et

                if not current_day:
                    continue

                # Track which group columns already got data this row (to avoid duplicates from merges)
                processed_cols = set()

                for col_idx, (sheet_name, gid, db_name) in block_groups.items():
                    if col_idx in processed_cols:
                        continue

                    # Use get_cell to resolve merged cells
                    cell_val = get_cell(row_idx, col_idx)
                    if not cell_val:
                        continue

                    raw_text = str(cell_val).strip()
                    if not raw_text or raw_text in ("-", "—", " "):
                        continue
                    # Skip formula cells (e.g., =COUNTIF(...))
                    if raw_text.startswith("="):
                        continue

                    # Check if this cell is from a merged range covering multiple group columns
                    # If so, create records for ALL covered groups
                    target_groups = [(col_idx, sheet_name, gid, db_name)]

                    # Find the origin cell of this merged range
                    origin_col = col_idx
                    for min_r, max_r, min_c, max_c, val in data_merge_ranges:
                        if min_r <= row_idx <= max_r and min_c <= col_idx <= max_c:
                            origin_col = min_c
                            break

                    coverage_key = (row_idx, origin_col)
                    if coverage_key in merged_data_coverage:
                        covered = merged_data_coverage[coverage_key]
                        target_groups = []
                        for gc in sorted(covered):
                            if gc in block_groups and gc not in processed_cols:
                                sn, gi, dn = block_groups[gc]
                                target_groups.append((gc, sn, gi, dn))

                    subject, stype, teacher, room, building = self._parse_cell_content(raw_text)

                    for tg_col, tg_sheet_name, tg_gid, tg_db_name in target_groups:
                        processed_cols.add(tg_col)
                        rec = {
                            "sheet_group_name": tg_sheet_name,
                            "db_group_name": tg_db_name,
                            "group_id": tg_gid,
                            "day": current_day,
                            "lesson_number": current_lesson or 1,
                            "subject": subject,
                            "teacher": teacher,
                            "room": room,
                            "building": building,
                            "start_time": current_start_time,
                            "end_time": current_end_time,
                            "schedule_type": stype or ScheduleType.LECTURE,
                            "_raw_cell": raw_text,
                        }
                        if subject or raw_text:
                            records.append(rec)

        return records

    # ══════════════════════════════════════════════════════
    # IMPORT: STUDENTS (simple)
    # ══════════════════════════════════════════════════════

    async def import_students(
        self,
        file_data: bytes,
        group_id: Optional[int] = None,
        update_existing: bool = True,
    ) -> Dict[str, Any]:
        """Import students from Excel — UPSERT by JSHSHIR or student_id."""
        try:
            df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl')
        except Exception as e:
            raise BadRequestException(f"Excel faylni o'qib bo'lmadi: {str(e)}")

        column_map = {
            "F.I.O": "name", "FIO": "name", "Ism": "name",
            "Telefon": "phone", "Tel": "phone",
            "Email": "email", "Pochta": "email",
            "Tug'ilgan sana": "birth_date", "Tug'ilgan": "birth_date",
            "Jinsi": "gender", "Jins": "gender",
            "Manzil": "address", "Pasport": "passport",
            "JSHSHIR": "jshshir", "PINFL": "jshshir",
            "Kontrakt": "contract_amount", "Shartnoma": "contract_amount",
            "To'langan": "contract_paid", "Guruh": "group_name",
        }
        df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

        total = len(df)
        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                name = str(row.get("name", "")).strip()
                if not name:
                    skipped += 1
                    continue

                student_data: Dict[str, Any] = {
                    "name": name,
                    "phone": str(row.get("phone", "")).strip() or None,
                    "email": str(row.get("email", "")).strip() or None,
                    "passport": str(row.get("passport", "")).strip() or None,
                    "jshshir": str(row.get("jshshir", "")).strip() or None,
                    "address": str(row.get("address", "")).strip() or None,
                    "group_id": group_id,
                }

                # Parse birth_date
                birth_date = row.get("birth_date")
                if pd.notna(birth_date):
                    if isinstance(birth_date, str):
                        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                            try:
                                student_data["birth_date"] = datetime.strptime(birth_date, fmt).date()
                                break
                            except ValueError:
                                pass
                    elif isinstance(birth_date, datetime):
                        student_data["birth_date"] = birth_date.date()

                # Parse gender
                gender = str(row.get("gender", "")).strip().lower()
                if gender in ["erkak", "male", "m"]:
                    student_data["gender"] = "male"
                elif gender in ["ayol", "female", "f"]:
                    student_data["gender"] = "female"

                # Parse contract
                contract = row.get("contract_amount", 0)
                if pd.notna(contract):
                    student_data["contract_amount"] = Decimal(str(contract))
                paid = row.get("contract_paid", 0)
                if pd.notna(paid):
                    student_data["contract_paid"] = Decimal(str(paid))

                # Resolve group
                group_name = str(row.get("group_name", "")).strip()
                if group_name and not group_id:
                    group_result = await self.db.execute(select(Group).where(Group.name == group_name))
                    group = group_result.scalar_one_or_none()
                    if group:
                        student_data["group_id"] = group.id

                # UPSERT by JSHSHIR
                existing = None
                if student_data.get("jshshir"):
                    result = await self.db.execute(
                        select(Student).where(Student.jshshir == student_data["jshshir"])
                    )
                    existing = result.scalar_one_or_none()

                if existing and update_existing:
                    for key, value in student_data.items():
                        if value is not None:
                            setattr(existing, key, value)
                    updated += 1
                elif not existing:
                    year = now_tashkent().year
                    prefix = f"ST-{year}-"
                    max_result = await self.db.execute(
                        select(Student.student_id)
                        .where(Student.student_id.like(f"{prefix}%"))
                        .order_by(Student.student_id.desc())
                        .limit(1)
                    )
                    max_id = max_result.scalar()
                    num = int(max_id.split("-")[-1]) + 1 if max_id else 1
                    student_data["student_id"] = f"{prefix}{num:04d}"

                    student = Student(**student_data)
                    self.db.add(student)
                    imported += 1
                else:
                    skipped += 1

            except Exception as e:
                failed += 1
                errors.append({"row": idx + 2, "name": str(row.get("name", "")), "error": str(e)})

        await self.db.commit()
        return {
            "success": True,
            "total": total,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "failed": failed,
            "errors": errors[:50],
        }

    # ══════════════════════════════════════════════════════
    # IMPORT: GROUPS
    # ══════════════════════════════════════════════════════

    async def import_groups(
        self,
        file_data: bytes,
        update_existing: bool = True,
    ) -> Dict[str, Any]:
        """Import groups from Excel — UPSERT by name."""
        try:
            df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl')
        except Exception as e:
            raise BadRequestException(f"Excel faylni o'qib bo'lmadi: {str(e)}")

        column_map = {
            "Guruh kodi": "name", "Guruh nomi": "name", "Guruh": "name",
            "Kod": "name", "Name": "name", "Nomi": "name",
            "Fakultet": "faculty", "Faculty": "faculty", "Yo'nalish": "faculty",
            "Kurs": "course_year", "Course": "course_year", "Bosqich": "course_year",
            "Kontrakt": "contract_amount", "Shartnoma summasi": "contract_amount",
        }
        df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

        total = len(df)
        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                name = str(row.get("name", "")).strip()
                if not name:
                    skipped += 1
                    continue

                faculty = str(row.get("faculty", "")).strip() or "Noma'lum"
                course_year = row.get("course_year", 1)
                course_year = 1 if pd.isna(course_year) else int(course_year)
                contract_amount = row.get("contract_amount", 0)
                contract_amount = Decimal("0") if pd.isna(contract_amount) else Decimal(str(contract_amount))

                result = await self.db.execute(select(Group).where(Group.name == name))
                existing = result.scalar_one_or_none()

                if existing:
                    if update_existing:
                        existing.faculty = faculty
                        existing.course_year = course_year
                        existing.contract_amount = contract_amount
                        updated += 1
                    else:
                        skipped += 1
                else:
                    group = Group(
                        name=name, faculty=faculty, course_year=course_year,
                        contract_amount=contract_amount, is_active=True,
                    )
                    self.db.add(group)
                    imported += 1

            except Exception as e:
                failed += 1
                errors.append({"row": idx + 2, "name": str(row.get("name", "")), "error": str(e)})

        await self.db.commit()
        return {
            "success": True, "total": total, "imported": imported,
            "updated": updated, "skipped": skipped, "failed": failed,
            "errors": errors[:50],
        }

    # ══════════════════════════════════════════════════════
    # IMPORT: ATTENDANCE
    # ══════════════════════════════════════════════════════

    async def import_attendance(
        self,
        file_data: bytes,
        group_id: int,
        attendance_date: date,
    ) -> Dict[str, Any]:
        """Import attendance from Excel — UPSERT by student + date."""
        try:
            df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl')
        except Exception as e:
            raise BadRequestException(f"Excel faylni o'qib bo'lmadi: {str(e)}")

        column_map = {
            "Talaba ID": "student_id", "ID": "student_id",
            "F.I.O": "name", "FIO": "name", "Ism": "name",
            "Holat": "status", "Status": "status",
            "Izoh": "note", "Note": "note", "Sabab": "note",
        }
        df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

        status_map = {
            "keldi": AttendanceStatus.PRESENT, "present": AttendanceStatus.PRESENT,
            "bor": AttendanceStatus.PRESENT, "+": AttendanceStatus.PRESENT,
            "kelmadi": AttendanceStatus.ABSENT, "absent": AttendanceStatus.ABSENT,
            "yo'q": AttendanceStatus.ABSENT, "-": AttendanceStatus.ABSENT,
            "kech": AttendanceStatus.LATE, "late": AttendanceStatus.LATE,
            "kech qoldi": AttendanceStatus.LATE,
            "sababli": AttendanceStatus.EXCUSED, "excused": AttendanceStatus.EXCUSED,
        }

        # Load group students
        students_result = await self.db.execute(
            select(Student).where(Student.group_id == group_id)
        )
        students = students_result.scalars().all()
        student_by_id = {s.student_id: s for s in students}
        student_by_name = {s.name.upper(): s for s in students}

        imported = 0
        failed = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                # Find student
                student = None
                sid = str(row.get("student_id", "")).strip()
                if sid and sid in student_by_id:
                    student = student_by_id[sid]
                else:
                    name = str(row.get("name", "")).strip().upper()
                    if name in student_by_name:
                        student = student_by_name[name]

                if not student:
                    failed += 1
                    errors.append({"row": idx + 2, "error": "Talaba topilmadi"})
                    continue

                status_str = str(row.get("status", "")).strip().lower()
                status = status_map.get(status_str, AttendanceStatus.PRESENT)
                note = str(row.get("note", "")).strip() or None

                # UPSERT attendance
                existing = await self.db.execute(
                    select(Attendance).where(
                        Attendance.student_id == student.id,
                        Attendance.date == attendance_date,
                    )
                )
                att = existing.scalar_one_or_none()

                if att:
                    att.status = status
                    att.note = note
                else:
                    att = Attendance(
                        student_id=student.id,
                        date=attendance_date,
                        status=status,
                        note=note,
                    )
                    self.db.add(att)

                imported += 1

            except Exception as e:
                failed += 1
                errors.append({"row": idx + 2, "error": str(e)})

        await self.db.commit()
        return {
            "success": True, "imported": imported, "failed": failed,
            "errors": errors[:50],
        }
