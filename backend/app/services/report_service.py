"""
UniControl - Report Service
===========================
Handles report generation, statistics, and management.

Author: UniControl Team
Version: 2.0.0
"""

from datetime import datetime, date
from app.config import now_tashkent, today_tashkent
from typing import Optional, List, Tuple
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from loguru import logger

from app.models.report import Report, ReportType, ReportFormat, ReportStatus
from app.models.student import Student
from app.models.attendance import Attendance, AttendanceStatus
from app.schemas.report import ReportCreate, ReportUpdate, ReportResponse
from app.core.exceptions import NotFoundException


class ReportService:
    """Report generation service."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def get_by_id(self, report_id: int) -> Optional[Report]:
        """Get report by ID."""
        result = await self.db.execute(
            select(Report).where(Report.id == report_id)
        )
        return result.scalar_one_or_none()
    
    async def list_reports(
        self,
        page: int = 1,
        page_size: int = 20,
        report_type: Optional[ReportType] = None,
        status: Optional[ReportStatus] = None,
        created_by: Optional[int] = None,
        group_id: Optional[int] = None,
    ) -> Tuple[List[Report], int]:
        """List reports with pagination and filters."""
        query = select(Report)
        count_query = select(func.count(Report.id))
        
        if report_type:
            query = query.where(Report.report_type == report_type)
            count_query = count_query.where(Report.report_type == report_type)
        
        if status:
            query = query.where(Report.status == status)
            count_query = count_query.where(Report.status == status)
        
        if created_by:
            query = query.where(Report.created_by == created_by)
            count_query = count_query.where(Report.created_by == created_by)
        
        if group_id:
            query = query.where(Report.group_id == group_id)
            count_query = count_query.where(Report.group_id == group_id)
        
        total_result = await self.db.execute(count_query)
        total = total_result.scalar()
        
        query = query.order_by(Report.created_at.desc())
        query = query.offset((page - 1) * page_size).limit(page_size)
        
        result = await self.db.execute(query)
        reports = result.scalars().all()
        
        return list(reports), total
    
    async def create(
        self,
        report_data: ReportCreate,
        user_id: int
    ) -> Report:
        """Create a report."""
        report = Report(
            name=report_data.name,
            description=report_data.description,
            report_type=report_data.report_type,
            format=report_data.format,
            date_from=report_data.date_from,
            date_to=report_data.date_to,
            group_id=report_data.group_id,
            filters=report_data.filters,
            created_by=user_id,
            status=ReportStatus.PENDING,
        )
        
        self.db.add(report)
        await self.db.commit()
        await self.db.refresh(report)
        
        return report
    
    async def process_report(self, report_id: int) -> Report:
        """Process a pending report."""
        report = await self.get_by_id(report_id)
        if not report:
            raise NotFoundException("Report not found")
        
        report.status = ReportStatus.PROCESSING
        report.started_at = now_tashkent()
        await self.db.commit()
        
        try:
            # Generate report based on type
            if report.report_type == ReportType.ATTENDANCE:
                file_path = await self._generate_attendance_report(report)
            elif report.report_type == ReportType.PAYMENT:
                file_path = await self._generate_payment_report(report)
            elif report.report_type == ReportType.STUDENTS:
                file_path = await self._generate_students_report(report)
            else:
                file_path = await self._generate_generic_report(report)
            
            report.file_path = file_path
            report.status = ReportStatus.COMPLETED
            report.completed_at = now_tashkent()
            
        except Exception as e:
            report.status = ReportStatus.FAILED
            report.error_message = str(e)
            report.completed_at = now_tashkent()
        
        await self.db.commit()
        await self.db.refresh(report)
        
        return report
    
    async def _generate_attendance_report(self, report: Report) -> str:
        """Generate attendance report data."""
        # This will be implemented with ExcelService
        return f"reports/attendance_{report.id}.xlsx"
    
    async def _generate_payment_report(self, report: Report) -> str:
        """Generate payment report data."""
        return f"reports/payment_{report.id}.xlsx"
    
    async def _generate_students_report(self, report: Report) -> str:
        """Generate students report data."""
        return f"reports/students_{report.id}.xlsx"
    
    async def _generate_generic_report(self, report: Report) -> str:
        """Generate generic report data."""
        return f"reports/report_{report.id}.xlsx"
    
    async def delete(self, report_id: int) -> bool:
        """Delete a report."""
        report = await self.get_by_id(report_id)
        if not report:
            raise NotFoundException("Report not found")
        
        await self.db.delete(report)
        await self.db.commit()
        logger.info(f"Report deleted: id={report_id}")
        
        return True

    async def update(self, report_id: int, report_data: ReportUpdate) -> Report:
        """Update a report."""
        report = await self.get_by_id(report_id)
        if not report:
            raise NotFoundException("Report not found")

        update_fields = report_data.model_dump(exclude_unset=True)
        for field, value in update_fields.items():
            setattr(report, field, value)

        report.updated_at = now_tashkent()
        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def update_status(
        self,
        report_id: int,
        new_status: ReportStatus,
        user_id: int,
        reason: Optional[str] = None,
    ) -> Report:
        """Update report status (approve/reject)."""
        report = await self.get_by_id(report_id)
        if not report:
            raise NotFoundException("Report not found")

        report.status = new_status
        report.approved_by = user_id
        report.approved_at = now_tashkent()
        if reason:
            report.rejection_reason = reason
        report.updated_at = now_tashkent()
        await self.db.commit()
        await self.db.refresh(report)
        logger.info(f"Report status updated: id={report_id} -> {new_status.value} by user={user_id}")
        return report

    async def generate_report(
        self,
        report_type: ReportType,
        group_id: Optional[int],
        start_date: Optional[date],
        end_date: Optional[date],
        created_by: int,
    ) -> Report:
        """Create and process a report."""
        report = Report(
            name=f"{report_type.value.replace('_', ' ').title()} Report",
            description=f"Auto-generated {report_type.value} report",
            report_type=report_type,
            format=ReportFormat.EXCEL,
            date_from=start_date,
            date_to=end_date,
            group_id=group_id,
            created_by=created_by,
            status=ReportStatus.PENDING,
        )
        self.db.add(report)
        await self.db.commit()
        await self.db.refresh(report)

        # Process the report
        return await self.process_report(report.id)

    async def download_report(
        self, report_id: int, format: str = "pdf"
    ) -> tuple:
        """
        Download report file.
        Returns (file_data, content_type, filename).
        """
        report = await self.get_by_id(report_id)
        if not report:
            raise NotFoundException("Report not found")

        # Increment download counter
        report.download_count = (report.download_count or 0) + 1
        await self.db.commit()

        report_name = (report.name or "report").replace(" ", "_")

        if format == "pdf":
            from app.services.pdf_service import PDFReportGenerator
            pdf_gen = PDFReportGenerator(self.db)
            file_data = await pdf_gen.generate_pdf(report)
            return (
                file_data,
                "application/pdf",
                f"{report_name}_{report_id}.pdf",
            )

        elif format == "excel":
            file_data = await self._generate_excel(report)
            return (
                file_data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{report_name}_{report_id}.xlsx",
            )

        elif format == "csv":
            file_data = await self._generate_csv(report)
            return file_data, "text/csv; charset=utf-8", f"{report_name}_{report_id}.csv"

        else:
            # Default to PDF
            from app.services.pdf_service import PDFReportGenerator
            pdf_gen = PDFReportGenerator(self.db)
            file_data = await pdf_gen.generate_pdf(report)
            return (
                file_data,
                "application/pdf",
                f"{report_name}_{report_id}.pdf",
            )

    async def _generate_excel(self, report) -> bytes:
        """Generate Excel report with real data."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Hisobot"

        # Styles
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        # Title row
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = report.name or "Hisobot"
        title_cell.font = Font(name='Arial', bold=True, size=16, color='6D28D9')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # Info rows
        ws['A3'] = "Turi:"
        ws['B3'] = report.report_type.value.replace("_", " ").title()
        ws['A4'] = "Holat:"
        ws['B4'] = report.status.value.upper()
        ws['A5'] = "Davr:"
        ws['B5'] = f"{report.date_from or '—'} — {report.date_to or '—'}"

        start_row = 7

        # Fetch data based on type
        if report.report_type.value in ("attendance",):
            headers = ["#", "Talaba", "Sana", "Holat", "Izoh"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            query = select(Attendance)
            if report.date_from:
                query = query.where(Attendance.date >= report.date_from)
            if report.date_to:
                query = query.where(Attendance.date <= report.date_to)
            if report.group_id:
                query = query.join(Student, Student.id == Attendance.student_id).where(
                    Student.group_id == report.group_id
                )
            result = await self.db.execute(query.order_by(Attendance.date.desc()))
            attendances = result.scalars().all()

            for idx, att in enumerate(attendances, 1):
                row = start_row + idx
                ws.cell(row=row, column=1, value=idx).border = thin_border
                # Try to get student name
                s_result = await self.db.execute(select(Student.name).where(Student.id == att.student_id))
                s_name = s_result.scalar() or f"ID: {att.student_id}"
                ws.cell(row=row, column=2, value=s_name).border = thin_border
                ws.cell(row=row, column=3, value=str(att.date)).border = thin_border
                status_map = {"present": "Keldi", "absent": "Kelmadi", "late": "Kechikdi"}
                ws.cell(row=row, column=4, value=status_map.get(att.status.value, att.status.value)).border = thin_border
                ws.cell(row=row, column=5, value=att.note if hasattr(att, 'note') and att.note else "").border = thin_border

        elif report.report_type.value in ("payment", "students"):
            headers = ["#", "Talaba", "ID", "Telefon", "Kontrakt", "To'langan", "Qoldiq"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            query = select(Student).where(Student.is_active == True)
            if report.group_id:
                query = query.where(Student.group_id == report.group_id)
            result = await self.db.execute(query.order_by(Student.name))
            students = result.scalars().all()

            for idx, s in enumerate(students, 1):
                row = start_row + idx
                amount = float(s.contract_amount or 0)
                paid = float(s.contract_paid or 0)
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=s.name).border = thin_border
                ws.cell(row=row, column=3, value=s.student_id or "—").border = thin_border
                ws.cell(row=row, column=4, value=s.phone or "—").border = thin_border
                ws.cell(row=row, column=5, value=amount).border = thin_border
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=6, value=paid).border = thin_border
                ws.cell(row=row, column=6).number_format = '#,##0'
                ws.cell(row=row, column=7, value=max(amount - paid, 0)).border = thin_border
                ws.cell(row=row, column=7).number_format = '#,##0'
        else:
            # Generic
            ws.cell(row=start_row, column=1, value="Ma'lumot").font = header_font
            ws.cell(row=start_row, column=1).fill = header_fill
            ws.cell(row=start_row + 1, column=1, value=report.description or "Maxsus hisobot")

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _generate_csv(self, report) -> bytes:
        """Generate CSV report with real data."""
        import csv as csv_module
        import io as io_module

        output = io_module.StringIO()
        writer = csv_module.writer(output)

        if report.report_type.value in ("attendance",):
            writer.writerow(["#", "Talaba", "Sana", "Holat"])
            query = select(Attendance)
            if report.date_from:
                query = query.where(Attendance.date >= report.date_from)
            if report.date_to:
                query = query.where(Attendance.date <= report.date_to)
            if report.group_id:
                query = query.join(Student, Student.id == Attendance.student_id).where(
                    Student.group_id == report.group_id
                )
            result = await self.db.execute(query.order_by(Attendance.date.desc()))
            attendances = result.scalars().all()
            for idx, att in enumerate(attendances, 1):
                s_result = await self.db.execute(select(Student.name).where(Student.id == att.student_id))
                s_name = s_result.scalar() or f"ID: {att.student_id}"
                writer.writerow([idx, s_name, str(att.date), att.status.value])
        else:
            writer.writerow(["#", "Talaba", "ID", "Kontrakt", "To'langan", "Qoldiq"])
            query = select(Student).where(Student.is_active == True)
            if report.group_id:
                query = query.where(Student.group_id == report.group_id)
            result = await self.db.execute(query.order_by(Student.name))
            students = result.scalars().all()
            for idx, s in enumerate(students, 1):
                amount = float(s.contract_amount or 0)
                paid = float(s.contract_paid or 0)
                writer.writerow([idx, s.name, s.student_id or "", amount, paid, max(amount - paid, 0)])

        return output.getvalue().encode("utf-8-sig")

    async def get_report_stats(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        created_by: Optional[int] = None,
        group_id: Optional[int] = None,
    ) -> dict:
        """
        Get report statistics summary.

        - If created_by is set — stats for that user's reports (leader view).
        - If group_id is set — stats for that group.
        - Otherwise — system-wide.
        """
        query = select(Report)
        if created_by:
            query = query.where(Report.created_by == created_by)
        if group_id:
            query = query.where(Report.group_id == group_id)
        if start_date:
            query = query.where(Report.created_at >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            query = query.where(Report.created_at <= datetime.combine(end_date, datetime.max.time()))

        result = await self.db.execute(query)
        reports = result.scalars().all()

        total = len(reports)
        pending = sum(1 for r in reports if r.status == ReportStatus.PENDING)
        approved = sum(1 for r in reports if r.status == ReportStatus.APPROVED)
        rejected = sum(1 for r in reports if r.status == ReportStatus.REJECTED)
        completed = sum(1 for r in reports if r.status == ReportStatus.COMPLETED)
        failed = sum(1 for r in reports if r.status == ReportStatus.FAILED)

        by_type = {}
        for r in reports:
            t = r.report_type.value
            by_type[t] = by_type.get(t, 0) + 1

        return {
            "total": total,
            "by_status": {
                "pending": pending,
                "approved": approved,
                "rejected": rejected,
                "completed": completed,
                "failed": failed,
            },
            "by_type": by_type,
            "approval_rate": round(approved / total * 100, 1) if total > 0 else 0,
        }
    
    async def get_dashboard_stats(
        self,
        group_id: Optional[int] = None
    ) -> dict:
        """Get dashboard statistics."""
        # Students stats
        student_query = select(func.count(Student.id)).where(Student.is_active == True)
        if group_id:
            student_query = student_query.where(Student.group_id == group_id)
        student_result = await self.db.execute(student_query)
        total_students = student_result.scalar()
        
        # Today's attendance
        today = today_tashkent()
        attendance_query = select(Attendance).where(Attendance.date == today)
        if group_id:
            attendance_query = attendance_query.join(Student).where(
                Student.group_id == group_id
            )
        attendance_result = await self.db.execute(attendance_query)
        today_attendance = attendance_result.scalars().all()
        
        present = sum(1 for a in today_attendance if a.status == AttendanceStatus.PRESENT)
        absent = sum(1 for a in today_attendance if a.status == AttendanceStatus.ABSENT)
        late = sum(1 for a in today_attendance if a.status == AttendanceStatus.LATE)
        
        # Payment stats
        payment_result = await self.db.execute(
            select(
                func.sum(Student.contract_amount),
                func.sum(Student.contract_paid)
            ).where(Student.is_active == True)
        )
        payment_data = payment_result.one()
        total_contract = float(payment_data[0] or 0)
        total_paid = float(payment_data[1] or 0)
        
        return {
            "total_students": total_students,
            "today_attendance": {
                "present": present,
                "absent": absent,
                "late": late,
                "rate": float(present / (present + absent + late) * 100) if (present + absent + late) > 0 else 0
            },
            "payment": {
                "total_contract": total_contract,
                "total_paid": total_paid,
                "remaining": total_contract - total_paid,
                "percentage": float(total_paid / total_contract * 100) if total_contract > 0 else 0
            }
        }