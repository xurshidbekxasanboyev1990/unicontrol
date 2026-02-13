"""
UniControl - Contract Service
===============================
Handles contract management operations including Excel import.

Author: UniControl Team
Version: 1.0.0
"""

import io
import logging
from decimal import Decimal, InvalidOperation
from typing import Optional, List, Tuple
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, delete
from sqlalchemy.orm import joinedload

from app.models.contract import Contract
from app.models.student import Student
from app.models.group import Group
from app.schemas.contract import (
    ContractCreate,
    ContractUpdate,
    ContractStats,
    ContractImportResult,
)
from app.core.exceptions import NotFoundException, ConflictException, BadRequestException

logger = logging.getLogger(__name__)


def safe_decimal(value, default=0) -> Decimal:
    """Safely convert a value to Decimal."""
    if value is None or (isinstance(value, str) and value.strip() in ('', ' ', '-')):
        return Decimal(str(default))
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(str(default))


def safe_float(value, default=None) -> Optional[float]:
    """Safely convert a value to float."""
    if value is None or (isinstance(value, str) and value.strip() in ('', ' ', '-')):
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_str(value) -> Optional[str]:
    """Safely convert a value to string."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s != ' ' else None


class ContractService:
    """Contract management service."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def get_by_id(self, contract_id: int) -> Optional[Contract]:
        """Get contract by ID with relationships."""
        result = await self.db.execute(
            select(Contract)
            .options(
                joinedload(Contract.student).joinedload(Student.group)
            )
            .where(Contract.id == contract_id)
        )
        return result.unique().scalar_one_or_none()
    
    async def get_by_student_and_year(
        self, student_id: int, academic_year: str
    ) -> Optional[Contract]:
        """Get contract by student and academic year."""
        result = await self.db.execute(
            select(Contract)
            .where(
                Contract.student_id == student_id,
                Contract.academic_year == academic_year
            )
        )
        return result.scalar_one_or_none()
    
    async def list_contracts(
        self,
        page: int = 1,
        page_size: int = 50,
        academic_year: Optional[str] = None,
        group_id: Optional[int] = None,
        education_form: Optional[str] = None,
        student_status: Optional[str] = None,
        has_debt: Optional[bool] = None,
        search: Optional[str] = None,
    ) -> Tuple[List[Contract], int]:
        """List contracts with pagination and filters."""
        query = (
            select(Contract)
            .options(
                joinedload(Contract.student).joinedload(Student.group)
            )
        )
        count_query = select(func.count(Contract.id))
        
        # Filters that need join
        if group_id or search:
            query = query.join(Contract.student)
            count_query = count_query.join(Contract.student)
        
        if academic_year:
            query = query.where(Contract.academic_year == academic_year)
            count_query = count_query.where(Contract.academic_year == academic_year)
        
        if group_id:
            query = query.where(Student.group_id == group_id)
            count_query = count_query.where(Student.group_id == group_id)
        
        if education_form:
            query = query.where(Contract.education_form == education_form)
            count_query = count_query.where(Contract.education_form == education_form)
        
        if student_status:
            query = query.where(Contract.student_status == student_status)
            count_query = count_query.where(Contract.student_status == student_status)
        
        if has_debt is True:
            query = query.where(Contract.debt_amount < 0)
            count_query = count_query.where(Contract.debt_amount < 0)
        elif has_debt is False:
            query = query.where(Contract.debt_amount >= 0)
            count_query = count_query.where(Contract.debt_amount >= 0)
        
        if search:
            search_filter = f"%{search}%"
            if not group_id:
                # Already joined above if group_id, otherwise need join
                pass
            query = query.where(
                or_(
                    Student.name.ilike(search_filter),
                    Student.jshshir.ilike(search_filter),
                    Student.passport.ilike(search_filter),
                    Student.phone.ilike(search_filter),
                )
            )
            count_query = count_query.where(
                or_(
                    Student.name.ilike(search_filter),
                    Student.jshshir.ilike(search_filter),
                    Student.passport.ilike(search_filter),
                    Student.phone.ilike(search_filter),
                )
            )
        
        total_result = await self.db.execute(count_query)
        total = total_result.scalar()
        
        query = query.order_by(Contract.id.desc())
        query = query.offset((page - 1) * page_size).limit(page_size)
        
        result = await self.db.execute(query)
        contracts = result.unique().scalars().all()
        
        return list(contracts), total
    
    async def create(self, data: ContractCreate) -> Contract:
        """Create a new contract."""
        # Check student exists
        student_result = await self.db.execute(
            select(Student).where(Student.id == data.student_id)
        )
        student = student_result.scalar_one_or_none()
        if not student:
            raise NotFoundException("Student not found")
        
        # Check uniqueness
        existing = await self.get_by_student_and_year(data.student_id, data.academic_year)
        if existing:
            raise ConflictException(
                f"Contract already exists for student {data.student_id} in {data.academic_year}"
            )
        
        contract = Contract(**data.model_dump())
        self.db.add(contract)
        await self.db.commit()
        await self.db.refresh(contract)
        
        return contract
    
    async def update(self, contract_id: int, data: ContractUpdate) -> Contract:
        """Update a contract."""
        contract = await self.get_by_id(contract_id)
        if not contract:
            raise NotFoundException("Contract not found")
        
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(contract, field, value)
        
        await self.db.commit()
        await self.db.refresh(contract)
        
        return contract
    
    async def delete(self, contract_id: int) -> bool:
        """Delete a contract."""
        contract = await self.get_by_id(contract_id)
        if not contract:
            raise NotFoundException("Contract not found")
        
        await self.db.delete(contract)
        await self.db.commit()
        return True
    
    async def delete_by_year(self, academic_year: str) -> int:
        """Delete all contracts for an academic year."""
        result = await self.db.execute(
            delete(Contract).where(Contract.academic_year == academic_year)
        )
        await self.db.commit()
        return result.rowcount
    
    async def get_statistics(
        self,
        academic_year: Optional[str] = None,
        group_id: Optional[int] = None,
    ) -> ContractStats:
        """Get contract statistics."""
        query = select(Contract).options(
            joinedload(Contract.student).joinedload(Student.group)
        )
        
        if academic_year:
            query = query.where(Contract.academic_year == academic_year)
        
        if group_id:
            query = query.join(Contract.student).where(Student.group_id == group_id)
        
        result = await self.db.execute(query)
        contracts = result.unique().scalars().all()
        
        total = len(contracts)
        total_contract = sum(c.contract_amount for c in contracts)
        total_paid = sum(c.total_paid for c in contracts)
        total_debt = sum(c.debt_amount for c in contracts)
        total_grant = sum(c.grant_amount for c in contracts)
        total_refund = sum(c.refund_amount for c in contracts)
        fully_paid = sum(1 for c in contracts if c.is_fully_paid)
        with_debt = sum(1 for c in contracts if c.debt_amount < 0)
        
        kunduzgi = sum(1 for c in contracts if c.education_form and 'kunduzgi' in c.education_form.lower())
        sirtqi = sum(1 for c in contracts if c.education_form and 'sirtqi' in c.education_form.lower())
        
        studying = sum(1 for c in contracts if c.student_status and "o'qimoqda" in c.student_status.lower())
        academic = sum(1 for c in contracts if c.student_status and "akademik" in c.student_status.lower())
        other_status = total - studying - academic
        
        return ContractStats(
            total_contracts=total,
            total_contract_amount=total_contract,
            total_paid=total_paid,
            total_debt=total_debt,
            total_grant_amount=total_grant,
            total_refund=total_refund,
            payment_percentage=float(total_paid / total_contract * 100) if total_contract > 0 else 0,
            fully_paid_count=fully_paid,
            with_debt_count=with_debt,
            kunduzgi_count=kunduzgi,
            sirtqi_count=sirtqi,
            studying_count=studying,
            academic_leave_count=academic,
            other_status_count=other_status,
        )
    
    async def get_academic_years(self) -> List[str]:
        """Get list of all academic years in contracts."""
        result = await self.db.execute(
            select(Contract.academic_year)
            .distinct()
            .order_by(Contract.academic_year.desc())
        )
        return [row[0] for row in result.all()]
    
    async def get_education_forms(self) -> List[str]:
        """Get list of all education forms."""
        result = await self.db.execute(
            select(Contract.education_form)
            .distinct()
            .where(Contract.education_form.isnot(None))
            .order_by(Contract.education_form)
        )
        return [row[0] for row in result.all()]
    
    async def get_student_statuses(self) -> List[str]:
        """Get list of all student statuses."""
        result = await self.db.execute(
            select(Contract.student_status)
            .distinct()
            .where(Contract.student_status.isnot(None))
            .order_by(Contract.student_status)
        )
        return [row[0] for row in result.all()]
    
    async def import_from_excel(
        self,
        file_data: bytes,
        academic_year: str = "2025-2026",
        update_existing: bool = True,
    ) -> ContractImportResult:
        """
        Import contract data from Excel file.
        
        Uses raw SQL UPSERT for reliability — handles duplicates in Excel,
        recovers from errors per-row, and uses batch commits.
        """
        import openpyxl
        from sqlalchemy import text
        
        errors = []
        warnings = []
        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        total_rows = 0
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active
        except Exception as e:
            return ContractImportResult(
                success=False,
                total_rows=0,
                imported=0,
                updated=0,
                skipped=0,
                failed=0,
                errors=[f"Excel faylni o'qishda xatolik: {str(e)}"],
            )
        
        # Pre-load all students for fast lookup (read-only, no ORM tracking)
        students_result = await self.db.execute(
            select(
                Student.id, Student.name, Student.jshshir, 
                Student.passport, Student.phone, Student.group_id
            )
        )
        all_students = students_result.all()
        
        # Pre-load groups
        groups_result = await self.db.execute(select(Group.id, Group.name))
        all_groups = groups_result.all()
        group_id_to_name = {g.id: g.name for g in all_groups}
        
        # Build lookup dictionaries (student_id -> student_row)
        jshshir_map = {}
        passport_map = {}
        name_group_map = {}
        
        for s in all_students:
            if s.jshshir:
                jshshir_map[s.jshshir.strip()] = s
            if s.passport:
                passport_map[s.passport.strip().upper()] = s
            # Name + group for fallback matching
            if s.group_id and s.group_id in group_id_to_name:
                gname = group_id_to_name[s.group_id]
                key = f"{s.name.strip().upper()}|{gname.strip().upper()}"
                name_group_map[key] = s
        
        # Also build group name aliases (DI-25-01 vs DI_25_01)
        group_name_map = {}
        for g in all_groups:
            n = g.name.strip().upper()
            group_name_map[n] = g
            group_name_map[n.replace('_', '-')] = g
            group_name_map[n.replace('-', '_')] = g
        
        # Detect data start row
        data_start = 4
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None and isinstance(cell_val, (int, float)):
                data_start = row_idx
                break
        
        logger.info(f"Excel import: sheet='{ws.title}', max_row={ws.max_row}, data_start={data_start}")
        
        # Track already-processed student IDs within this import (handle Excel duplicates)
        seen_student_ids = set()
        
        # Pre-load existing contracts for this academic year
        existing_contracts_result = await self.db.execute(
            select(Contract.student_id).where(Contract.academic_year == academic_year)
        )
        existing_contract_student_ids = {r[0] for r in existing_contracts_result.all()}
        
        # Collect rows for batch processing
        upsert_data = []
        student_updates = []
        
        for row_idx in range(data_start, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=2).value
            if not name_val or not str(name_val).strip():
                continue
            
            total_rows += 1
            name = str(name_val).strip()
            
            try:
                # Read all columns
                jshshir_raw = ws.cell(row=row_idx, column=3).value
                jshshir = None
                if jshshir_raw is not None:
                    try:
                        jshshir = str(int(float(str(jshshir_raw).strip())))
                    except (ValueError, TypeError):
                        jshshir = str(jshshir_raw).strip()
                
                passport_raw = ws.cell(row=row_idx, column=4).value
                passport = safe_str(passport_raw)
                if passport:
                    passport = passport.upper()
                
                phone_raw = ws.cell(row=row_idx, column=5).value
                phone = safe_str(phone_raw)
                if phone:
                    try:
                        phone = str(int(float(phone)))
                    except (ValueError, TypeError):
                        pass
                
                course = safe_str(ws.cell(row=row_idx, column=6).value)
                student_status = safe_str(ws.cell(row=row_idx, column=7).value)
                group_name = safe_str(ws.cell(row=row_idx, column=8).value)
                direction = safe_str(ws.cell(row=row_idx, column=9).value)
                education_form = safe_str(ws.cell(row=row_idx, column=10).value)
                
                contract_amount = safe_decimal(ws.cell(row=row_idx, column=11).value, 0)
                grant_percentage = safe_float(ws.cell(row=row_idx, column=12).value)
                grant_amount = safe_decimal(ws.cell(row=row_idx, column=13).value, 0)
                debt_amount = safe_decimal(ws.cell(row=row_idx, column=14).value, 0)
                payment_percentage = safe_float(ws.cell(row=row_idx, column=15).value)
                total_paid = safe_decimal(ws.cell(row=row_idx, column=16).value, 0)
                refund_amount = safe_decimal(ws.cell(row=row_idx, column=17).value, 0)
                year_start_balance = safe_decimal(ws.cell(row=row_idx, column=18).value, 0)
                year_end_balance = safe_decimal(ws.cell(row=row_idx, column=19).value, 0)
                
                # Find student: JSHSHIR → passport → name+group
                student = None
                
                if jshshir and jshshir in jshshir_map:
                    student = jshshir_map[jshshir]
                elif passport and passport in passport_map:
                    student = passport_map[passport]
                elif group_name:
                    key = f"{name.upper()}|{group_name.upper()}"
                    if key in name_group_map:
                        student = name_group_map[key]
                
                if not student:
                    skipped += 1
                    if total_rows <= 100 or skipped <= 50:
                        warnings.append(
                            f"Qator {row_idx}: '{name}' topilmadi (JSHSHIR={jshshir})"
                        )
                    continue
                
                sid = student.id
                
                # Handle Excel duplicates — if we already processed this student, update it
                if sid in seen_student_ids:
                    # This is a duplicate in Excel — treat as update
                    for i, d in enumerate(upsert_data):
                        if d['student_id'] == sid:
                            upsert_data[i] = {
                                'student_id': sid,
                                'academic_year': academic_year,
                                'course': course,
                                'student_status': student_status,
                                'direction': direction,
                                'education_form': education_form,
                                'contract_amount': float(contract_amount),
                                'grant_percentage': grant_percentage,
                                'grant_amount': float(grant_amount),
                                'debt_amount': float(debt_amount),
                                'payment_percentage': payment_percentage,
                                'total_paid': float(total_paid),
                                'refund_amount': float(refund_amount),
                                'year_start_balance': float(year_start_balance),
                                'year_end_balance': float(year_end_balance),
                            }
                            break
                    warnings.append(f"Qator {row_idx}: '{name}' Excel'da takrorlangan — yangilandi")
                    continue
                
                seen_student_ids.add(sid)
                
                is_existing = sid in existing_contract_student_ids
                if is_existing and not update_existing:
                    skipped += 1
                    continue
                
                if is_existing:
                    updated += 1
                else:
                    imported += 1
                
                upsert_data.append({
                    'student_id': sid,
                    'academic_year': academic_year,
                    'course': course,
                    'student_status': student_status,
                    'direction': direction,
                    'education_form': education_form,
                    'contract_amount': float(contract_amount),
                    'grant_percentage': grant_percentage,
                    'grant_amount': float(grant_amount),
                    'debt_amount': float(debt_amount),
                    'payment_percentage': payment_percentage,
                    'total_paid': float(total_paid),
                    'refund_amount': float(refund_amount),
                    'year_start_balance': float(year_start_balance),
                    'year_end_balance': float(year_end_balance),
                })
                
                # Collect student updates
                student_updates.append({
                    'id': sid,
                    'contract_amount': float(contract_amount),
                    'contract_paid': float(total_paid),
                    'phone': phone if phone and not student.phone else None,
                    'passport': passport if passport and not student.passport else None,
                })
                
            except Exception as e:
                failed += 1
                if failed <= 50:
                    errors.append(f"Qator {row_idx} ({name}): {str(e)}")
                logger.error(f"Contract import parse error at row {row_idx}: {e}")
        
        logger.info(
            f"Excel parsed: total_rows={total_rows}, to_upsert={len(upsert_data)}, "
            f"skipped={skipped}, failed={failed}"
        )
        
        # Now execute batch UPSERT using raw SQL
        if upsert_data:
            try:
                # Process in batches of 200
                batch_size = 200
                for batch_start in range(0, len(upsert_data), batch_size):
                    batch = upsert_data[batch_start:batch_start + batch_size]
                    
                    for row in batch:
                        upsert_sql = text("""
                            INSERT INTO contracts (
                                student_id, academic_year, course, student_status,
                                direction, education_form, contract_amount,
                                grant_percentage, grant_amount, debt_amount,
                                payment_percentage, total_paid, refund_amount,
                                year_start_balance, year_end_balance,
                                created_at, updated_at
                            ) VALUES (
                                :student_id, :academic_year, :course, :student_status,
                                :direction, :education_form, :contract_amount,
                                :grant_percentage, :grant_amount, :debt_amount,
                                :payment_percentage, :total_paid, :refund_amount,
                                :year_start_balance, :year_end_balance,
                                NOW(), NOW()
                            )
                            ON CONFLICT (student_id, academic_year) DO UPDATE SET
                                course = EXCLUDED.course,
                                student_status = EXCLUDED.student_status,
                                direction = EXCLUDED.direction,
                                education_form = EXCLUDED.education_form,
                                contract_amount = EXCLUDED.contract_amount,
                                grant_percentage = EXCLUDED.grant_percentage,
                                grant_amount = EXCLUDED.grant_amount,
                                debt_amount = EXCLUDED.debt_amount,
                                payment_percentage = EXCLUDED.payment_percentage,
                                total_paid = EXCLUDED.total_paid,
                                refund_amount = EXCLUDED.refund_amount,
                                year_start_balance = EXCLUDED.year_start_balance,
                                year_end_balance = EXCLUDED.year_end_balance,
                                updated_at = NOW()
                        """)
                        await self.db.execute(upsert_sql, row)
                    
                    await self.db.commit()
                    logger.info(f"Contract import batch committed: {batch_start + len(batch)}/{len(upsert_data)}")
                
                # Update student contract fields
                for su in student_updates:
                    update_parts = ["contract_amount = :contract_amount", "contract_paid = :contract_paid"]
                    params = {
                        'id': su['id'],
                        'contract_amount': su['contract_amount'],
                        'contract_paid': su['contract_paid'],
                    }
                    if su.get('phone'):
                        update_parts.append("phone = :phone")
                        params['phone'] = su['phone']
                    if su.get('passport'):
                        update_parts.append("passport = :passport")
                        params['passport'] = su['passport']
                    
                    sql = text(f"UPDATE students SET {', '.join(update_parts)} WHERE id = :id")
                    await self.db.execute(sql, params)
                
                await self.db.commit()
                logger.info(f"Student contract fields updated: {len(student_updates)} students")
                
            except Exception as e:
                await self.db.rollback()
                logger.error(f"Contract import batch error: {e}", exc_info=True)
                errors.append(f"Ma'lumotlarni saqlashda xatolik: {str(e)}")
                return ContractImportResult(
                    success=False,
                    total_rows=total_rows,
                    imported=0,
                    updated=0,
                    skipped=skipped,
                    failed=total_rows - skipped,
                    errors=errors[:50],
                    warnings=warnings[:50],
                )
        
        logger.info(
            f"Contract import complete: total={total_rows}, imported={imported}, "
            f"updated={updated}, skipped={skipped}, failed={failed}"
        )
        
        return ContractImportResult(
            success=imported > 0 or updated > 0,
            total_rows=total_rows,
            imported=imported,
            updated=updated,
            skipped=skipped,
            failed=failed,
            errors=errors[:50],
            warnings=warnings[:100],
        )
    
    async def export_to_excel(
        self,
        academic_year: Optional[str] = None,
        group_id: Optional[int] = None,
    ) -> bytes:
        """Export contracts to Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        query = (
            select(Contract)
            .options(
                joinedload(Contract.student).joinedload(Student.group)
            )
        )
        
        if academic_year:
            query = query.where(Contract.academic_year == academic_year)
        if group_id:
            query = query.join(Contract.student).where(Student.group_id == group_id)
        
        query = query.order_by(Contract.id)
        
        result = await self.db.execute(query)
        contracts = result.unique().scalars().all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = academic_year or "Kontrakt ma'lumotlari"
        
        # Header style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # Headers (matching Excel template)
        headers = [
            "№",
            "To'liq ismi",
            "JSHSHIR-kod",
            "Pasport seriya va raqami",
            "Telefon raqami",
            "Kurs",
            "Talaba xolati",
            "Guruh",
            "Yo'nalishi",
            "Ta'lim shakli",
            "Kontrakt summasi",
            "Grand (Foizda)",
            "Grand (Summada)",
            "Qarzdorlik summasi",
            "Foizda (Jami)",
            "To'lov summasi",
            "Qaytarilgan summa",
            "Yil boshiga qoldiq",
            "Yil yakuniga qoldiq",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for idx, contract in enumerate(contracts, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=contract.student_name or "").border = thin_border
            ws.cell(row=row, column=3, value=contract.student_jshshir or "").border = thin_border
            ws.cell(row=row, column=4, value=contract.student.passport if contract.student else "").border = thin_border
            ws.cell(row=row, column=5, value=contract.student.phone if contract.student else "").border = thin_border
            ws.cell(row=row, column=6, value=contract.course or "").border = thin_border
            ws.cell(row=row, column=7, value=contract.student_status or "").border = thin_border
            ws.cell(row=row, column=8, value=contract.group_name or "").border = thin_border
            ws.cell(row=row, column=9, value=contract.direction or "").border = thin_border
            ws.cell(row=row, column=10, value=contract.education_form or "").border = thin_border
            ws.cell(row=row, column=11, value=float(contract.contract_amount)).border = thin_border
            ws.cell(row=row, column=12, value=contract.grant_percentage or 0).border = thin_border
            ws.cell(row=row, column=13, value=float(contract.grant_amount)).border = thin_border
            ws.cell(row=row, column=14, value=float(contract.debt_amount)).border = thin_border
            ws.cell(row=row, column=15, value=contract.payment_percentage or 0).border = thin_border
            ws.cell(row=row, column=16, value=float(contract.total_paid)).border = thin_border
            ws.cell(row=row, column=17, value=float(contract.refund_amount)).border = thin_border
            ws.cell(row=row, column=18, value=float(contract.year_start_balance)).border = thin_border
            ws.cell(row=row, column=19, value=float(contract.year_end_balance)).border = thin_border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
