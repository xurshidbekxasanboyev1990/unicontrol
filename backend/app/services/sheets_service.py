"""
UniControl - Google Sheets (Excel) Service
=============================================
Reads university schedule data from an Excel (.xlsx) file hosted on Google Drive.

The file is an .xlsx (not a native Google Sheet) so we:
  1. Download it via the Google Drive API
  2. Parse it in-memory with openpyxl
  3. Write back via the Sheets API v4 REST (works on .xlsx too)

Spreadsheet structure (per sheet/tab):
  Row 1: Faculty/direction title (merged cells, value usually in D1+)
  Row 2: Group names in columns D, E, F, G, ... (e.g., "FTO'(ing) 25-01")
  Row 3+: Schedule data grouped by weekday
    Column A: Day name (Dushanba, Seshanba, etc.) - merged vertically
    Column B: Lesson number (1-5)
    Column C: Time range (e.g., "08:30-09:50")
    Column D+: Subject info per group (subject name + teacher + room)

Author: UniControl Team
Version: 2.0.0 — rewritten from gspread to openpyxl + Drive API
"""

import io
import re
import logging
import time as _time
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, time
from loguru import logger as loguru_logger

import requests as http_requests
import openpyxl
from google.oauth2.service_account import Credentials
from google.auth.transport.requests import Request as GoogleAuthRequest

from app.config import settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

WEEKDAY_MAP = {
    "dushanba": "monday",
    "seshanba": "tuesday",
    "chorshanba": "wednesday",
    "payshanba": "thursday",
    "juma": "friday",
    "shanba": "saturday",
    "yakshanba": "sunday",
}

TIME_SLOTS = {
    1: ("08:30", "09:50"),
    2: ("10:00", "11:20"),
    3: ("12:00", "13:20"),
    4: ("13:30", "14:50"),
    5: ("15:00", "16:20"),
    6: ("16:30", "17:50"),
}

DRIVE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/drive",
]


def _normalize_key(name: str) -> str:
    """
    Create a minimal canonical key from a group name for fuzzy matching.
    Strips apostrophes, normalizes separators, uppercases.
    
    Examples:
      "FTO'(ing) 25-01"   -> "FTO(ING)_25-01"
      "FTO'(ing)_25-01"   -> "FTO(ING)_25-01"
      "FTO' (o'zb)_25-01" -> "FTO(OZB)_25-01"
      "FTO'(o'zb) 25-01"  -> "FTO(OZB)_25-01"
      "KI-25-09"           -> "KI_25-09"
      "KI_25-09"           -> "KI_25-09"
      "IQ-25-01"           -> "IQ_25-01"
      "IQ_25-34 (rus)"     -> "IQ_25-34"
    """
    if not name:
        return ""
    
    text = name.strip().upper()
    
    # Remove all types of apostrophes
    for ch in ("'", "\u2018", "\u2019", "`", "\u02BB", "\u02BC"):
        text = text.replace(ch, "")
    
    # Remove trailing language tags like " (RUS)", " (UZB)" that appear as suffixes
    # These are annotations, not part of the group code
    # But KEEP parenthetical parts that are INSIDE the group code like "(ING)" "(OZB)" "(RUS)"
    # Pattern: group code ends with digits, then optional space + (lang) or just bare lang word
    text = re.sub(r'(\d{2})\s*\((?:rus|uzb|eng)\)$', r'\1', text, flags=re.IGNORECASE)
    text = re.sub(r'(\d{2})\s+(?:RUS|UZB|ENG|INGLIZ)$', r'\1', text, flags=re.IGNORECASE)
    
    # Collapse space before parenthesis: "FTO (OZB)" -> "FTO(OZB)"
    text = re.sub(r'\s+\(', '(', text)
    
    # Collapse spaces inside parentheses: "( OZB)" -> "(OZB)"
    text = re.sub(r'\(\s+', '(', text)
    text = re.sub(r'\s+\)', ')', text)
    
    # Normalize separator between prefix/paren and number part:
    # "FTO(ING) 25-01" -> "FTO(ING)_25-01"
    # "KI-25-09" -> "KI_25-09"  
    # Match: (letters or closing paren) then (space/hyphen/underscore) then (2 digits)
    text = re.sub(r'([A-Z\)])\s*[-_\s]\s*(\d{2})', r'\1_\2', text)
    
    # Also handle case with no separator: "FTO(ING)25-01" -> "FTO(ING)_25-01"
    text = re.sub(r'([A-Z\)])(\d{2})', r'\1_\2', text)
    
    # Collapse multiple separators
    text = re.sub(r'[_\s]{2,}', '_', text)
    
    return text.strip()


# Known prefix aliases: sheet prefix -> possible DB prefixes
# These handle cases where the abbreviation differs between sheets and DB
PREFIX_ALIASES = {
    "MM": ["MMT", "MM", "M"],       # Sheet: MM-25-01, DB: MMT_25-01
    "MMT": ["MMT", "MM", "M"],
    "M": ["M", "MM", "MMT"],
    "TR": ["TR"],
    "PS": ["PS"],
    "KI": ["KI"],
    "IQ": ["IQ"],
    "BT": ["BT"],
    "MT": ["MT"],
}

# Explicit overrides for groups that cannot be auto-matched
EXPLICIT_MAP = {
    "Magistr_FTO'(ing) 25-01": "Mag_Ingliz-25_01",
}


def build_group_name_map(sheet_group_names: List[str], db_group_names: List[str]) -> Dict[str, str]:
    """
    Build a mapping from sheet group names to DB group names.
    
    Uses multi-tier matching:
      1. Exact match
      2. Normalized key match (handles apostrophe/separator differences)
      3. Prefix alias match (handles MM -> MMT, etc.)
      4. Fuzzy fallback (stripped alphanumeric comparison)
    
    Returns: {sheet_name: db_name} for matched groups.
    """
    # Pre-compute normalized keys for all DB groups
    db_by_key = {}       # normalized_key -> db_name
    db_by_stripped = {}   # stripped (only alnum) -> db_name
    
    for db_name in db_group_names:
        key = _normalize_key(db_name)
        db_by_key[key] = db_name
        
        # Also store exact name as key
        db_by_key[db_name] = db_name
        db_by_key[db_name.upper()] = db_name
        
        # Stripped version: only letters + digits
        stripped = re.sub(r'[^A-Z0-9]', '', db_name.upper())
        db_by_stripped[stripped] = db_name
    
    result: Dict[str, str] = {}
    db_names_set = set(db_group_names)

    for sheet_name in sheet_group_names:
        # --- Tier 0: Explicit override ---
        if sheet_name in EXPLICIT_MAP and EXPLICIT_MAP[sheet_name] in db_names_set:
            result[sheet_name] = EXPLICIT_MAP[sheet_name]
            continue

        # --- Tier 1: Exact match ---
        if sheet_name in db_by_key:
            result[sheet_name] = db_by_key[sheet_name]
            continue
        
        # --- Tier 2: Normalized key match ---
        sheet_key = _normalize_key(sheet_name)
        if sheet_key in db_by_key:
            result[sheet_name] = db_by_key[sheet_key]
            continue
        
        # --- Tier 3: Prefix alias match ---
        # Extract prefix and number part from sheet_key
        # e.g. "MM_25-01" -> prefix="MM", suffix="25-01"
        alias_match = re.match(r'^([A-Z]+)(?:\([^)]*\))?_(.+)$', sheet_key)
        if alias_match:
            sheet_prefix = alias_match.group(1)
            sheet_suffix = alias_match.group(2)  # e.g. "25-01"
            
            # Try each alias prefix
            aliases = PREFIX_ALIASES.get(sheet_prefix, [])
            for alt_prefix in aliases:
                if alt_prefix == sheet_prefix:
                    continue  # already tried
                # Try simple: ALT_SUFFIX
                alt_key = f"{alt_prefix}_{sheet_suffix}"
                if alt_key in db_by_key:
                    result[sheet_name] = db_by_key[alt_key]
                    break
            
            if sheet_name in result:
                continue
        
        # --- Tier 4: Stripped alphanumeric match ---
        sheet_stripped = re.sub(r'[^A-Z0-9]', '', sheet_name.upper())
        if sheet_stripped in db_by_stripped:
            result[sheet_name] = db_by_stripped[sheet_stripped]
            continue
        
        # --- Tier 5: Smart partial match ---
        # Extract just the core: prefix + year + number, ignoring language tags
        # "FTO(ING)_25-01" -> try matching "FTO_25-01" variations
        core_match = re.match(r'^([A-Z]+)\([^)]*\)_(.+)$', sheet_key)
        if core_match:
            core_prefix = core_match.group(1)
            core_suffix = core_match.group(2)
            
            # Try: PREFIX(...)_SUFFIX with different paren content
            for db_key, db_name in db_by_key.items():
                db_key_upper = db_key.upper() if isinstance(db_key, str) else ""
                # Match same prefix + same numbers but different paren content
                db_core = re.match(r'^([A-Z]+)(?:\s*\([^)]*\))?\s*_(.+)$', db_key_upper)
                if db_core and db_core.group(1) == core_prefix and db_core.group(2) == core_suffix:
                    result[sheet_name] = db_name
                    break
    
    return result


def _parse_cell_content(value: str) -> Optional[Dict[str, Any]]:
    """
    Parse a schedule cell content into structured data.
    Example: "Jismoniy tarbiya va sport (amaliy) To'xtanazarov Qaxramon Sport zal B bino"
    Example: "Kelajak soati Faollar zali A bino"
    Returns: {subject, type, teacher, room, building}
    """
    if not value or not str(value).strip():
        return None

    text = str(value).strip()

    # Extract type (ma'ruza, amaliy, seminar, laboratoriya)
    schedule_type = "lecture"
    type_match = re.search(r'\((ma\'?ruza|amaliy|seminar|laboratoriya)\)', text, re.IGNORECASE)
    if type_match:
        type_str = type_match.group(1).lower()
        if "amaliy" in type_str:
            schedule_type = "practice"
        elif "seminar" in type_str:
            schedule_type = "seminar"
        elif "laboratoriya" in type_str:
            schedule_type = "lab"
        else:
            schedule_type = "lecture"

    # Extract room and building
    room = None
    building = None
    room_match = re.search(r'(\d{3})-?xona?\s*(A|B)\s*bino', text, re.IGNORECASE)
    if room_match:
        room = f"{room_match.group(1)}-xona"
        building = f"{room_match.group(2)} bino"
    else:
        # Try other patterns like "Sport zal B bino", "L-1 B bino", "Faollar zali A bino"
        special_match = re.search(r'(Sport\s*zal(?:i)?|L-\d+|Faollar\s*zali|Laboratoriya|Akt\s*zal(?:i)?|Majlis\s*zali)\s*(A|B)?\s*bino', text, re.IGNORECASE)
        if special_match:
            room = special_match.group(1).strip()
            building = f"{special_match.group(2) or 'A'} bino"
            room_match = special_match  # reuse for subject cleanup below

    # Extract subject (everything before the type marker OR before room)
    subject = text
    if type_match:
        subject = text[:type_match.start()].strip()
    elif room_match:
        # No type marker — strip room/building from the end of the text
        subject = text[:room_match.start()].strip()

    # Extract teacher name (after type marker, before room)
    teacher = None
    if type_match:
        after_type = text[type_match.end():].strip()
        # Remove room/building info
        if room_match:
            after_type = text[type_match.end():room_match.start()].strip()
        elif "Faollar zali" in after_type:
            after_type = after_type.split("Faollar")[0].strip()
        elif "Sport zal" in after_type:
            after_type = after_type.split("Sport")[0].strip()

        if after_type:
            teacher = after_type.strip()

    return {
        "subject": subject or text,
        "type": schedule_type,
        "teacher": teacher,
        "room": room,
        "building": building,
        "raw": text,
    }


# ---------------------------------------------------------------------------
# Drive download helper
# ---------------------------------------------------------------------------

_credentials_cache: Optional[Credentials] = None
_workbook_cache: Dict[str, Any] = {"wb": None, "ts": 0}
CACHE_TTL = 120  # seconds – re-download every 2 minutes max


def _get_drive_credentials() -> Credentials:
    """Get (and cache) Google service-account credentials."""
    global _credentials_cache
    if _credentials_cache is None or not _credentials_cache.valid:
        _credentials_cache = Credentials.from_service_account_file(
            settings.GOOGLE_SHEETS_CREDENTIALS_PATH,
            scopes=DRIVE_SCOPES,
        )
        _credentials_cache.refresh(GoogleAuthRequest())
    elif _credentials_cache.expired:
        _credentials_cache.refresh(GoogleAuthRequest())
    return _credentials_cache


def _download_workbook(force: bool = False) -> openpyxl.Workbook:
    """
    Download the .xlsx from Google Drive and return an openpyxl Workbook.
    Results are cached for CACHE_TTL seconds.
    """
    now = _time.time()
    if (
        not force
        and _workbook_cache["wb"] is not None
        and (now - _workbook_cache["ts"]) < CACHE_TTL
    ):
        return _workbook_cache["wb"]

    creds = _get_drive_credentials()
    if not creds.token:
        creds.refresh(GoogleAuthRequest())

    file_id = settings.GOOGLE_SHEETS_SPREADSHEET_ID
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {creds.token}"}

    logger.info(f"Downloading .xlsx from Google Drive (file_id={file_id}) ...")
    resp = http_requests.get(url, headers=headers, timeout=60)

    if resp.status_code != 200:
        logger.error(f"Drive download failed: {resp.status_code} {resp.text[:300]}")
        raise RuntimeError(f"Failed to download Excel file from Drive: {resp.status_code}")

    wb = openpyxl.load_workbook(
        io.BytesIO(resp.content),
        data_only=True,
    )
    _workbook_cache["wb"] = wb
    _workbook_cache["ts"] = now
    logger.info(f"Excel loaded: {len(wb.sheetnames)} sheets, {len(resp.content)} bytes")
    return wb


def _build_merge_map(ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Build a map from every cell inside a merged range to the top-left anchor cell.
    Key: (row, col) 1-based  ->  Value: (anchor_row, anchor_col) 1-based
    """
    merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        anchor = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != anchor:
                    merge_map[(r, c)] = anchor
    return merge_map


def _ws_to_rows(ws) -> List[List[str]]:
    """
    Read an openpyxl worksheet into a list of string lists ('' for None).
    Merged cells are expanded: every cell in a merged range gets the anchor value.
    """
    merge_map = _build_merge_map(ws)
    # Pre-read all cell values into a dict for fast lookup
    cell_values: Dict[Tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_values[(cell.row, cell.column)] = str(cell.value).strip()

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    rows: List[List[str]] = []
    for r in range(1, max_row + 1):
        row_data: List[str] = []
        for c in range(1, max_col + 1):
            if (r, c) in merge_map:
                # This cell is part of a merged range – use anchor's value
                anchor = merge_map[(r, c)]
                row_data.append(cell_values.get(anchor, ""))
            else:
                row_data.append(cell_values.get((r, c), ""))
        rows.append(row_data)
    return rows


def _build_merge_ranges(ws) -> List[Dict[str, Any]]:
    """
    Return a list of merge range descriptors for the DATA area (col >= 4).
    Each entry: {min_row, max_row, min_col, max_col, value}
    """
    ranges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col >= 4 and mr.max_col > mr.min_col:
            val = ws.cell(mr.min_row, mr.min_col).value
            ranges.append({
                "min_row": mr.min_row,
                "max_row": mr.max_row,
                "min_col": mr.min_col,
                "max_col": mr.max_col,
                "value": str(val).strip() if val else "",
            })
    return ranges


# ---------------------------------------------------------------------------
# Write-back via Google Sheets API v4 (REST)
# ---------------------------------------------------------------------------

def _col_to_letter(c: int) -> str:
    """Convert 1-based column number to A1 letter(s). 1->A, 26->Z, 27->AA."""
    s = ""
    while c > 0:
        c, remainder = divmod(c - 1, 26)
        s = chr(65 + remainder) + s
    return s


def _sheets_api_update_cell(sheet_title: str, row: int, col: int, value: str) -> Dict[str, Any]:
    """Update a single cell via the Google Sheets API v4 (REST)."""
    creds = _get_drive_credentials()
    if not creds.token:
        creds.refresh(GoogleAuthRequest())

    file_id = settings.GOOGLE_SHEETS_SPREADSHEET_ID
    cell_ref = f"{_col_to_letter(col)}{row}"
    range_str = f"'{sheet_title}'!{cell_ref}"

    url = (
        f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}"
        f"/values/{http_requests.utils.quote(range_str, safe='')}"
        f"?valueInputOption=USER_ENTERED"
    )
    headers = {
        "Authorization": f"Bearer {creds.token}",
        "Content-Type": "application/json",
    }
    body = {"range": range_str, "majorDimension": "ROWS", "values": [[value]]}

    resp = http_requests.put(url, headers=headers, json=body, timeout=30)
    if resp.status_code != 200:
        logger.error(f"Sheets API update failed: {resp.status_code} {resp.text[:300]}")
        raise RuntimeError(f"Failed to update cell: {resp.status_code} — {resp.text[:200]}")

    # Invalidate cache
    _workbook_cache["wb"] = None
    _workbook_cache["ts"] = 0

    return {"success": True, "row": row, "col": col, "value": value}


def _sheets_api_batch_update(sheet_title: str, updates: List[Dict]) -> Dict[str, Any]:
    """Batch update multiple cells via the Google Sheets API v4 (REST)."""
    creds = _get_drive_credentials()
    if not creds.token:
        creds.refresh(GoogleAuthRequest())

    file_id = settings.GOOGLE_SHEETS_SPREADSHEET_ID
    data_entries = []
    for u in updates:
        cell_ref = f"{_col_to_letter(u['col'])}{u['row']}"
        range_str = f"'{sheet_title}'!{cell_ref}"
        data_entries.append({
            "range": range_str,
            "majorDimension": "ROWS",
            "values": [[u["value"]]],
        })

    url = (
        f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}"
        f"/values:batchUpdate"
    )
    headers = {
        "Authorization": f"Bearer {creds.token}",
        "Content-Type": "application/json",
    }
    body = {"valueInputOption": "USER_ENTERED", "data": data_entries}

    resp = http_requests.post(url, headers=headers, json=body, timeout=60)
    if resp.status_code != 200:
        logger.error(f"Sheets API batch update failed: {resp.status_code} {resp.text[:300]}")
        raise RuntimeError(f"Batch update failed: {resp.status_code}")

    _workbook_cache["wb"] = None
    _workbook_cache["ts"] = 0

    return {"success": True, "updated": len(updates)}


# ---------------------------------------------------------------------------
# Main service class
# ---------------------------------------------------------------------------

class SheetsService:
    """Excel schedule management service (Drive download + openpyxl)."""

    def __init__(self):
        self.file_id = settings.GOOGLE_SHEETS_SPREADSHEET_ID

    # ------------------------------------------------------------------
    # Sheet listing
    # ------------------------------------------------------------------

    def get_sheet_names(self) -> List[Dict[str, Any]]:
        """Get all sheet (tab) names with basic info."""
        wb = _download_workbook()
        result = []
        for idx, name in enumerate(wb.sheetnames):
            ws = wb[name]
            result.append({
                "id": idx,
                "title": name,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "index": idx,
            })
        return result

    # ------------------------------------------------------------------
    # Parsed schedule data
    # ------------------------------------------------------------------

    def get_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get full parsed schedule data from a specific sheet.
        Returns faculty info, groups list, and schedule matrix.
        """
        wb = _download_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        all_values = _ws_to_rows(ws)

        if not all_values or len(all_values) < 3:
            return {"title": sheet_name, "groups": [], "schedule": []}

        # Row 1: Faculty title
        faculty_title = ""
        for cell in all_values[0]:
            if cell:
                faculty_title = cell
                break

        # Row 2: Group names (column D+ = index 3+)
        groups = []
        group_columns: Dict[int, str] = {}
        if len(all_values) > 1:
            for col_idx in range(3, len(all_values[1])):
                val = all_values[1][col_idx]
                if val:
                    groups.append({"index": col_idx, "name": val})
                    group_columns[col_idx] = val

        # Parse schedule rows (row 3+)
        schedule = []
        current_day = None

        for row_idx in range(2, len(all_values)):
            row = all_values[row_idx]
            if not row:
                continue

            # Column A: Day name
            day_val = row[0] if len(row) > 0 else ""
            if day_val:
                day_lower = day_val.lower().strip()
                if day_lower in WEEKDAY_MAP:
                    current_day = day_val

            if not current_day:
                continue

            # Column B: Lesson number
            lesson_num = None
            if len(row) > 1 and row[1]:
                try:
                    lesson_num = int(float(row[1]))
                except (ValueError, TypeError):
                    continue

            if not lesson_num:
                continue

            # Column C: Time range
            time_range = row[2] if len(row) > 2 else ""
            if not time_range and lesson_num in TIME_SLOTS:
                time_range = f"{TIME_SLOTS[lesson_num][0]}-{TIME_SLOTS[lesson_num][1]}"

            # Parse each group's cell
            lessons: Dict[str, Any] = {}
            for col_idx, group_name in group_columns.items():
                cell_val = row[col_idx] if col_idx < len(row) else ""
                if cell_val:
                    parsed = _parse_cell_content(cell_val)
                    if parsed:
                        lessons[group_name] = parsed

            if lessons:
                schedule.append({
                    "row": row_idx + 1,
                    "day": current_day,
                    "day_en": WEEKDAY_MAP.get(current_day.lower().strip(), ""),
                    "lesson_number": lesson_num,
                    "time": time_range,
                    "lessons": lessons,
                })

        return {
            "title": faculty_title,
            "sheet_name": sheet_name,
            "groups": groups,
            "groups_count": len(groups),
            "schedule": schedule,
            "total_lessons": sum(len(s["lessons"]) for s in schedule),
        }

    # ------------------------------------------------------------------
    # Raw data (2-D table for display)
    # ------------------------------------------------------------------

    def get_sheet_raw(self, sheet_name: str, max_rows: int = 100) -> Dict[str, Any]:
        """Get raw sheet data as a 2D string array for table display."""
        wb = _download_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        all_values = _ws_to_rows(ws)
        rows = all_values[:max_rows] if max_rows else all_values

        max_col = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                if row[i]:
                    if i + 1 > max_col:
                        max_col = i + 1
                    break

        trimmed = [row[:max_col] for row in rows]

        return {
            "sheet_name": sheet_name,
            "rows": trimmed,
            "total_rows": len(all_values),
            "total_cols": max_col,
            "displayed_rows": len(trimmed),
        }

    # ------------------------------------------------------------------
    # Cell editing
    # ------------------------------------------------------------------

    def update_cell(self, sheet_name: str, row: int, col: int, value: str) -> Dict[str, Any]:
        """Update a single cell."""
        return _sheets_api_update_cell(sheet_name, row, col, value)

    def update_cells_batch(self, sheet_name: str, updates: List[Dict]) -> Dict[str, Any]:
        """Batch update multiple cells."""
        return _sheets_api_batch_update(sheet_name, updates)

    # ------------------------------------------------------------------
    # Group schedule
    # ------------------------------------------------------------------

    def get_group_schedule(self, sheet_name: str, group_name: str) -> Dict[str, Any]:
        """Get schedule for a specific group from a sheet."""
        data = self.get_sheet_data(sheet_name)

        group_schedule = []
        for entry in data["schedule"]:
            if group_name in entry["lessons"]:
                lesson = entry["lessons"][group_name]
                group_schedule.append({
                    "day": entry["day"],
                    "day_en": entry["day_en"],
                    "lesson_number": entry["lesson_number"],
                    "time": entry["time"],
                    **lesson,
                })

        return {
            "group": group_name,
            "sheet_name": sheet_name,
            "faculty": data["title"],
            "schedule": group_schedule,
            "total_lessons": len(group_schedule),
        }

    # ------------------------------------------------------------------
    # Sync to database
    # ------------------------------------------------------------------

    def sync_to_database(self, sheet_name: str) -> Dict[str, Any]:
        """Parse sheet data → structured records for DB import."""
        data = self.get_sheet_data(sheet_name)
        records = []

        for entry in data["schedule"]:
            day_en = entry["day_en"]
            lesson_num = entry["lesson_number"]
            time_parts = entry["time"].split("-") if entry["time"] else []

            start_time = end_time = None
            if len(time_parts) == 2:
                try:
                    st = time_parts[0].strip()
                    et = time_parts[1].strip()
                    start_time = f"{st}:00" if len(st.split(":")) == 2 else st
                    end_time = f"{et}:00" if len(et.split(":")) == 2 else et
                except Exception as e:
                    loguru_logger.warning(f"Failed to parse time parts: {time_parts}, error: {e}")

            for group_name, lesson in entry["lessons"].items():
                records.append({
                    "group_name": group_name,
                    "subject": lesson["subject"],
                    "schedule_type": lesson["type"],
                    "teacher_name": lesson["teacher"],
                    "room": lesson["room"],
                    "building": lesson["building"],
                    "day_of_week": day_en,
                    "lesson_number": lesson_num,
                    "start_time": start_time,
                    "end_time": end_time,
                    "raw_text": lesson["raw"],
                })

        return {
            "sheet_name": sheet_name,
            "faculty": data["title"],
            "records": records,
            "total_records": len(records),
            "groups_affected": list(set(r["group_name"] for r in records)),
        }

    def sync_to_database_with_mapping(
        self, sheet_name: str, db_group_names: List[str]
    ) -> Dict[str, Any]:
        """Parse sheet data, build group name mapping, return records with resolved DB names."""
        sync_data = self.sync_to_database(sheet_name)
        records = sync_data["records"]

        sheet_groups = list(set(r["group_name"] for r in records))
        name_map = build_group_name_map(sheet_groups, db_group_names)

        for rec in records:
            sheet_name_val = rec["group_name"]
            rec["sheet_group_name"] = sheet_name_val
            rec["db_group_name"] = name_map.get(sheet_name_val)

        matched = {sg: dg for sg, dg in name_map.items()}
        unmatched = [sg for sg in sheet_groups if sg not in name_map]

        return {
            **sync_data,
            "records": records,
            "group_name_map": matched,
            "unmatched_groups": unmatched,
        }

    # ------------------------------------------------------------------
    # Full summary (all sheets)
    # ------------------------------------------------------------------

    def get_all_sheets_summary(self) -> Dict[str, Any]:
        """Get summary of all sheets with group counts."""
        sheets_info = self.get_sheet_names()
        summary = []

        for sheet_info in sheets_info:
            try:
                data = self.get_sheet_data(sheet_info["title"])
                summary.append({
                    "title": sheet_info["title"],
                    "faculty": data["title"],
                    "groups_count": data["groups_count"],
                    "groups": [g["name"] for g in data["groups"]],
                    "total_lessons": data["total_lessons"],
                })
            except Exception as e:
                summary.append({
                    "title": sheet_info["title"],
                    "faculty": sheet_info["title"],
                    "groups_count": 0,
                    "groups": [],
                    "total_lessons": 0,
                    "error": str(e),
                })

        return {
            "spreadsheet_id": self.file_id,
            "total_sheets": len(sheets_info),
            "sheets": summary,
            "total_groups": sum(s["groups_count"] for s in summary),
            "total_lessons": sum(s["total_lessons"] for s in summary),
        }
