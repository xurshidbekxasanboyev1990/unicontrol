"""
UniControl - Academic Affairs Panel API
========================================
Full schedule management for academic affairs department.
Includes CRUD, bulk operations, AI generation, group management.

Author: UniControl Team
Version: 1.0.0
"""

from typing import Optional, List
from datetime import time
from fastapi import APIRouter, Depends, Query, HTTPException, status, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, distinct, delete
from sqlalchemy.orm import joinedload
from pydantic import BaseModel

from app.database import get_db
from app.models.user import User, UserRole
from app.models.schedule import Schedule, WeekDay, ScheduleType, WeekType
from app.models.group import Group
from app.models.student import Student
from app.models.teacher_workload import TeacherWorkload
from app.core.dependencies import get_current_active_user, require_role
from app.services.claude_ai import generate_schedule_with_ai, ai_optimize_schedule

router = APIRouter()


# ============================================
# Dependencies
# ============================================

async def require_academic(
    current_user: User = Depends(get_current_active_user)
) -> User:
    """Require academic_affairs, admin, or superadmin role."""
    if current_user.role not in [UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.ACADEMIC_AFFAIRS]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Akademik ishlar paneli uchun ruxsat yo'q"
        )
    return current_user


# ============================================
# Pydantic Schemas
# ============================================

class ScheduleCreateRequest(BaseModel):
    group_id: int
    subject: str
    subject_code: Optional[str] = None
    schedule_type: str = "lecture"
    day_of_week: str
    start_time: str  # HH:MM
    end_time: str  # HH:MM
    lesson_number: Optional[int] = None
    week_type: Optional[str] = "all"
    room: Optional[str] = None
    building: Optional[str] = None
    teacher_name: Optional[str] = None
    teacher_id: Optional[int] = None
    color: Optional[str] = None
    semester: Optional[int] = None
    academic_year: Optional[str] = None

class ScheduleUpdateRequest(BaseModel):
    subject: Optional[str] = None
    subject_code: Optional[str] = None
    schedule_type: Optional[str] = None
    day_of_week: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    lesson_number: Optional[int] = None
    week_type: Optional[str] = None
    room: Optional[str] = None
    building: Optional[str] = None
    teacher_name: Optional[str] = None
    teacher_id: Optional[int] = None
    color: Optional[str] = None
    is_active: Optional[bool] = None
    is_cancelled: Optional[bool] = None

class BulkScheduleCreate(BaseModel):
    schedules: List[ScheduleCreateRequest]

class LectureMultiGroupRequest(BaseModel):
    """Create a lecture schedule for multiple groups at once."""
    group_ids: List[int]  # 2-5 groups to link
    subject: str
    subject_code: Optional[str] = None
    schedule_type: str = "lecture"
    day_of_week: str
    start_time: str
    end_time: str
    lesson_number: Optional[int] = None
    room: Optional[str] = None
    building: Optional[str] = None
    teacher_name: Optional[str] = None
    teacher_id: Optional[int] = None
    color: Optional[str] = None
    semester: Optional[int] = None
    academic_year: Optional[str] = None

class SubjectItem(BaseModel):
    name: str
    type: str = "lecture"
    hours_per_week: int = 2

class AIGenerateRequest(BaseModel):
    group_ids: List[int]
    subjects: List[SubjectItem]
    rooms: Optional[List[str]] = None
    constraints: Optional[str] = None
    language: str = "uz"

class AIOptimizeRequest(BaseModel):
    group_id: Optional[int] = None
    problem: str
    language: str = "uz"


# ============================================
# Helpers
# ============================================

def parse_time(t_str: str) -> time:
    """Parse time string HH:MM to time object."""
    parts = t_str.strip().split(":")
    return time(int(parts[0]), int(parts[1]))


async def resolve_teacher_id(
    db: AsyncSession,
    teacher_name: Optional[str] = None,
    teacher_id: Optional[int] = None
) -> tuple:
    """
    Auto-resolve teacher_id from teacher_name if not provided.
    Returns (teacher_id, teacher_name) tuple.
    
    Logic:
    1. If teacher_id is provided and valid, use it and fetch name
    2. If only teacher_name is provided, search for matching teacher in users
    3. Return resolved (teacher_id, teacher_name)
    """
    if teacher_id:
        # Verify teacher exists
        result = await db.execute(
            select(User).where(
                and_(User.id == teacher_id, User.role == UserRole.TEACHER, User.is_active == True)
            )
        )
        teacher = result.scalar_one_or_none()
        if teacher:
            return teacher.id, teacher.name
        # teacher_id invalid, fall through to name search
    
    if teacher_name and teacher_name.strip():
        name = teacher_name.strip()
        # Try exact match first
        result = await db.execute(
            select(User).where(
                and_(
                    User.role == UserRole.TEACHER,
                    User.is_active == True,
                    User.name.ilike(name)
                )
            )
        )
        teacher = result.scalar_one_or_none()
        if teacher:
            return teacher.id, teacher.name
        
        # Try partial match (contains)
        result = await db.execute(
            select(User).where(
                and_(
                    User.role == UserRole.TEACHER,
                    User.is_active == True,
                    User.name.ilike(f"%{name}%")
                )
            )
        )
        teachers = result.scalars().all()
        if len(teachers) == 1:
            return teachers[0].id, teachers[0].name
    
    return teacher_id, teacher_name

def schedule_to_dict(s: Schedule) -> dict:
    """Convert schedule to dict."""
    return {
        "id": s.id,
        "group_id": s.group_id,
        "group_name": s.group_name,
        "subject": s.subject,
        "subject_code": s.subject_code,
        "schedule_type": s.schedule_type.value if s.schedule_type else "lecture",
        "day_of_week": s.day_of_week.value if s.day_of_week else None,
        "start_time": s.start_time.strftime("%H:%M") if s.start_time else None,
        "end_time": s.end_time.strftime("%H:%M") if s.end_time else None,
        "time_range": s.time_range,
        "lesson_number": s.lesson_number,
        "week_type": s.week_type.value if s.week_type else "all",
        "room": s.room,
        "building": s.building,
        "teacher_name": s.teacher_name,
        "teacher_id": s.teacher_id,
        "color": s.color,
        "semester": s.semester,
        "academic_year": s.academic_year,
        "is_active": s.is_active,
        "is_cancelled": s.is_cancelled,
    }


# ============================================
# DASHBOARD
# ============================================

@router.get("/dashboard")
async def academic_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Academic panel dashboard statistics."""
    
    total_groups = (await db.execute(select(func.count(Group.id)).where(Group.is_active == True))).scalar() or 0
    total_schedules = (await db.execute(select(func.count(Schedule.id)).where(Schedule.is_active == True))).scalar() or 0
    total_subjects = (await db.execute(select(func.count(distinct(Schedule.subject))).where(Schedule.is_active == True))).scalar() or 0
    total_teachers = (await db.execute(
        select(func.count(User.id)).where(and_(User.role == UserRole.TEACHER, User.is_active == True))
    )).scalar() or 0
    
    # Groups without schedule
    groups_with_schedule = (await db.execute(
        select(distinct(Schedule.group_id)).where(Schedule.is_active == True)
    )).scalars().all()
    groups_without = (await db.execute(
        select(func.count(Group.id)).where(
            and_(Group.is_active == True, ~Group.id.in_(groups_with_schedule) if groups_with_schedule else True)
        )
    )).scalar() or 0
    
    return {
        "total_groups": total_groups,
        "total_schedules": total_schedules,
        "total_subjects": total_subjects,
        "total_teachers": total_teachers,
        "groups_without_schedule": groups_without,
    }


# ============================================
# GROUPS LIST (for schedule management)
# ============================================

@router.get("/groups")
async def academic_groups(
    faculty: Optional[str] = None,
    course_year: Optional[int] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get all groups with schedule statistics."""
    query = select(Group).where(Group.is_active == True)
    
    if faculty:
        query = query.where(Group.faculty == faculty)
    if course_year:
        query = query.where(Group.course_year == course_year)
    if search:
        query = query.where(Group.name.ilike(f"%{search}%"))
    
    query = query.order_by(Group.name)
    result = await db.execute(query)
    groups = result.scalars().all()
    
    items = []
    for g in groups:
        schedule_count = (await db.execute(
            select(func.count(Schedule.id)).where(
                and_(Schedule.group_id == g.id, Schedule.is_active == True)
            )
        )).scalar() or 0
        
        student_count = (await db.execute(
            select(func.count(Student.id)).where(
                and_(Student.group_id == g.id, Student.is_active == True)
            )
        )).scalar() or 0
        
        items.append({
            "id": g.id,
            "name": g.name,
            "faculty": g.faculty,
            "course_year": g.course_year,
            "students_count": student_count,
            "schedule_count": schedule_count,
            "is_active": g.is_active,
        })
    
    return {"items": items, "total": len(items)}


# ============================================
# SCHEDULE CRUD
# ============================================

@router.get("/schedules")
async def academic_list_schedules(
    group_id: Optional[int] = None,
    day_of_week: Optional[str] = None,
    semester: Optional[int] = None,
    academic_year: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """List all schedules with filters."""
    query = select(Schedule).options(joinedload(Schedule.group)).where(Schedule.is_active == True)
    
    if group_id:
        query = query.where(Schedule.group_id == group_id)
    if day_of_week:
        try:
            query = query.where(Schedule.day_of_week == WeekDay(day_of_week))
        except ValueError:
            pass
    if semester:
        query = query.where(Schedule.semester == semester)
    if academic_year:
        query = query.where(Schedule.academic_year == academic_year)
    
    query = query.order_by(Schedule.day_of_week, Schedule.start_time)
    result = await db.execute(query)
    schedules = result.unique().scalars().all()
    
    return {
        "items": [schedule_to_dict(s) for s in schedules],
        "total": len(schedules)
    }


@router.get("/schedules/group/{group_id}")
async def academic_group_schedule(
    group_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get full weekly schedule for a specific group (spreadsheet format)."""
    result = await db.execute(
        select(Schedule)
        .options(joinedload(Schedule.group))
        .where(and_(Schedule.group_id == group_id, Schedule.is_active == True))
        .order_by(Schedule.day_of_week, Schedule.start_time)
    )
    schedules = result.unique().scalars().all()
    
    # Group info
    group = (await db.execute(select(Group).where(Group.id == group_id))).scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    
    # Organize by day and lesson_number for spreadsheet
    grid = {}
    for s in schedules:
        day = s.day_of_week.value if s.day_of_week else "monday"
        ln = s.lesson_number or 0
        key = f"{day}_{ln}"
        grid[key] = schedule_to_dict(s)
    
    return {
        "group": {"id": group.id, "name": group.name, "faculty": group.faculty, "course_year": group.course_year},
        "items": [schedule_to_dict(s) for s in schedules],
        "grid": grid,
    }


@router.get("/schedules/multi-group")
async def academic_multi_group_schedule(
    group_ids: str = Query(..., description="Comma-separated group IDs"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get schedules for multiple groups (spreadsheet view with groups as columns)."""
    ids = [int(x.strip()) for x in group_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return {"groups": [], "items": [], "grid": {}}
    
    # Get groups
    groups_result = await db.execute(
        select(Group).where(Group.id.in_(ids)).order_by(Group.name)
    )
    groups = groups_result.scalars().all()
    
    # Get schedules
    result = await db.execute(
        select(Schedule)
        .options(joinedload(Schedule.group))
        .where(and_(Schedule.group_id.in_(ids), Schedule.is_active == True))
        .order_by(Schedule.day_of_week, Schedule.start_time)
    )
    schedules = result.unique().scalars().all()
    
    # Build grid: key = "day_lessonNumber_groupId"
    grid = {}
    for s in schedules:
        day = s.day_of_week.value if s.day_of_week else "monday"
        ln = s.lesson_number or 0
        key = f"{day}_{ln}_{s.group_id}"
        grid[key] = schedule_to_dict(s)
    
    return {
        "groups": [{"id": g.id, "name": g.name, "faculty": g.faculty} for g in groups],
        "items": [schedule_to_dict(s) for s in schedules],
        "grid": grid,
    }


@router.post("/schedules")
async def academic_create_schedule(
    data: ScheduleCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Create a single schedule entry."""
    try:
        schedule_type = ScheduleType(data.schedule_type)
    except ValueError:
        schedule_type = ScheduleType.LECTURE
    
    try:
        day = WeekDay(data.day_of_week)
    except ValueError:
        raise HTTPException(status_code=400, detail="Noto'g'ri kun")
    
    # Auto-resolve teacher_id from teacher_name
    resolved_teacher_id, resolved_teacher_name = await resolve_teacher_id(
        db, data.teacher_name, data.teacher_id
    )
    
    # Resolve week_type
    try:
        week_type = WeekType(data.week_type) if data.week_type else WeekType.ALL
    except ValueError:
        week_type = WeekType.ALL
    
    schedule = Schedule(
        group_id=data.group_id,
        subject=data.subject,
        subject_code=data.subject_code,
        schedule_type=schedule_type,
        day_of_week=day,
        start_time=parse_time(data.start_time),
        end_time=parse_time(data.end_time),
        lesson_number=data.lesson_number,
        week_type=week_type,
        room=data.room,
        building=data.building,
        teacher_name=resolved_teacher_name,
        teacher_id=resolved_teacher_id,
        color=data.color,
        semester=data.semester,
        academic_year=data.academic_year,
        is_active=True,
    )
    db.add(schedule)
    await db.commit()
    await db.refresh(schedule)
    
    return {"success": True, "id": schedule.id, "teacher_id": resolved_teacher_id, "message": "Jadval qo'shildi"}


@router.post("/schedules/lecture-multi")
async def academic_create_lecture_multi_group(
    data: LectureMultiGroupRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """
    Create a lecture schedule for multiple groups at once.
    When a lecture is added, it can be linked to 2-5 groups simultaneously.
    The same lesson is automatically copied to all selected groups.
    """
    if len(data.group_ids) < 1 or len(data.group_ids) > 5:
        raise HTTPException(status_code=400, detail="1 dan 5 gacha guruh tanlash mumkin")
    
    # Validate groups exist
    groups_result = await db.execute(
        select(Group).where(and_(Group.id.in_(data.group_ids), Group.is_active == True))
    )
    found_groups = groups_result.scalars().all()
    if len(found_groups) != len(data.group_ids):
        raise HTTPException(status_code=400, detail="Ba'zi guruhlar topilmadi")
    
    try:
        schedule_type = ScheduleType(data.schedule_type)
    except ValueError:
        schedule_type = ScheduleType.LECTURE
    
    try:
        day = WeekDay(data.day_of_week)
    except ValueError:
        raise HTTPException(status_code=400, detail="Noto'g'ri kun")
    
    # Auto-resolve teacher_id from teacher_name
    resolved_teacher_id, resolved_teacher_name = await resolve_teacher_id(
        db, data.teacher_name, data.teacher_id
    )
    
    created_ids = []
    group_names = []
    
    for group in found_groups:
        schedule = Schedule(
            group_id=group.id,
            subject=data.subject,
            subject_code=data.subject_code,
            schedule_type=schedule_type,
            day_of_week=day,
            start_time=parse_time(data.start_time),
            end_time=parse_time(data.end_time),
            lesson_number=data.lesson_number,
            week_type=WeekType.ALL,
            room=data.room,
            building=data.building,
            teacher_name=resolved_teacher_name,
            teacher_id=resolved_teacher_id,
            color=data.color,
            semester=data.semester,
            academic_year=data.academic_year,
            is_active=True,
        )
        db.add(schedule)
        created_ids.append(group.id)
        group_names.append(group.name)
    
    await db.commit()
    
    return {
        "success": True,
        "created_count": len(created_ids),
        "group_names": group_names,
        "message": f"{len(created_ids)} ta guruhga jadval qo'shildi: {', '.join(group_names)}"
    }


@router.post("/schedules/bulk")
async def academic_bulk_create(
    data: BulkScheduleCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Bulk create schedules (for AI-generated or copy operations)."""
    created = 0
    errors = []
    
    # Pre-load all teachers for efficient matching
    teachers_result = await db.execute(
        select(User).where(and_(User.role == UserRole.TEACHER, User.is_active == True))
    )
    all_teachers = teachers_result.scalars().all()
    teacher_name_map = {t.name.lower(): t for t in all_teachers}
    
    def quick_resolve_teacher(t_name, t_id):
        """Quick teacher resolution using pre-loaded cache."""
        if t_id:
            for t in all_teachers:
                if t.id == t_id:
                    return t.id, t.name
        if t_name and t_name.strip():
            name_lower = t_name.strip().lower()
            # Exact match
            if name_lower in teacher_name_map:
                t = teacher_name_map[name_lower]
                return t.id, t.name
            # Partial match
            matches = [t for key, t in teacher_name_map.items() if name_lower in key or key in name_lower]
            if len(matches) == 1:
                return matches[0].id, matches[0].name
        return t_id, t_name
    
    for idx, item in enumerate(data.schedules):
        try:
            try:
                schedule_type = ScheduleType(item.schedule_type)
            except ValueError:
                schedule_type = ScheduleType.LECTURE
            try:
                day = WeekDay(item.day_of_week)
            except ValueError:
                errors.append(f"#{idx+1}: Noto'g'ri kun - {item.day_of_week}")
                continue
            
            # Auto-resolve teacher
            resolved_id, resolved_name = quick_resolve_teacher(item.teacher_name, item.teacher_id)
            
            # Resolve week_type for item
            try:
                item_week_type = WeekType(item.week_type) if item.week_type else WeekType.ALL
            except ValueError:
                item_week_type = WeekType.ALL
            
            schedule = Schedule(
                group_id=item.group_id,
                subject=item.subject,
                subject_code=item.subject_code,
                schedule_type=schedule_type,
                day_of_week=day,
                start_time=parse_time(item.start_time),
                end_time=parse_time(item.end_time),
                lesson_number=item.lesson_number,
                week_type=item_week_type,
                room=item.room,
                building=item.building,
                teacher_name=resolved_name,
                teacher_id=resolved_id,
                color=item.color,
                semester=item.semester,
                academic_year=item.academic_year,
                is_active=True,
            )
            db.add(schedule)
            created += 1
        except Exception as e:
            errors.append(f"#{idx+1}: {str(e)}")
    
    await db.commit()
    
    return {
        "success": True,
        "created": created,
        "errors": errors,
        "message": f"{created} ta jadval qo'shildi" + (f", {len(errors)} ta xato" if errors else ""),
    }


@router.put("/schedules/{schedule_id}")
async def academic_update_schedule(
    schedule_id: int,
    data: ScheduleUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Update a schedule entry."""
    result = await db.execute(select(Schedule).where(Schedule.id == schedule_id))
    schedule = result.scalar_one_or_none()
    if not schedule:
        raise HTTPException(status_code=404, detail="Jadval topilmadi")
    
    update_data = data.model_dump(exclude_unset=True, exclude_none=True)
    
    for field, value in update_data.items():
        if field == "schedule_type":
            try:
                value = ScheduleType(value)
            except ValueError:
                continue
        elif field == "day_of_week":
            try:
                value = WeekDay(value)
            except ValueError:
                continue
        elif field == "week_type":
            try:
                value = WeekType(value)
            except ValueError:
                continue
        elif field in ("start_time", "end_time"):
            value = parse_time(value)
        setattr(schedule, field, value)
    
    # Auto-resolve teacher_id if teacher_name was updated
    if "teacher_name" in update_data or "teacher_id" in update_data:
        resolved_id, resolved_name = await resolve_teacher_id(
            db, schedule.teacher_name, schedule.teacher_id
        )
        schedule.teacher_id = resolved_id
        schedule.teacher_name = resolved_name
    
    await db.commit()
    return {"success": True, "message": "Jadval yangilandi"}


@router.delete("/schedules/{schedule_id}")
async def academic_delete_schedule(
    schedule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Delete a schedule entry."""
    result = await db.execute(select(Schedule).where(Schedule.id == schedule_id))
    schedule = result.scalar_one_or_none()
    if not schedule:
        raise HTTPException(status_code=404, detail="Jadval topilmadi")
    
    await db.delete(schedule)
    await db.commit()
    return {"success": True, "message": "Jadval o'chirildi"}


@router.delete("/schedules/group/{group_id}")
async def academic_clear_group_schedule(
    group_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Delete all schedules for a group."""
    await db.execute(
        delete(Schedule).where(Schedule.group_id == group_id)
    )
    await db.commit()
    return {"success": True, "message": "Guruh jadvali tozalandi"}


# ============================================
# TEACHERS LIST
# ============================================

@router.get("/teachers")
async def academic_teachers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get list of teachers for schedule assignment."""
    result = await db.execute(
        select(User)
        .where(and_(User.role == UserRole.TEACHER, User.is_active == True))
        .order_by(User.name)
    )
    teachers = result.scalars().all()
    
    return {
        "items": [
            {"id": t.id, "name": t.name, "login": t.login, "phone": t.phone, "email": t.email}
            for t in teachers
        ],
        "total": len(teachers)
    }


# ============================================
# SUBJECTS LIST (distinct from schedules)
# ============================================

@router.get("/subjects")
async def academic_subjects(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get list of unique subjects from all schedules."""
    result = await db.execute(
        select(distinct(Schedule.subject)).where(Schedule.is_active == True).order_by(Schedule.subject)
    )
    subjects = [row[0] for row in result.all()]
    return {"items": subjects, "total": len(subjects)}


# ============================================
# FACULTIES LIST
# ============================================

@router.get("/faculties")
async def academic_faculties(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get list of unique faculties."""
    result = await db.execute(
        select(distinct(Group.faculty)).where(Group.is_active == True).order_by(Group.faculty)
    )
    faculties = [row[0] for row in result.all()]
    return {"items": faculties, "total": len(faculties)}


# ============================================
# FACULTY HIERARCHY (Fakultet → Yo'nalish → Guruh)
# ============================================

# Mapping: direction (Group.faculty) → super-faculty
FACULTY_MAPPING = {
    # IT va Texnika fakulteti
    "Kompyuter injiniringi": "IT va Texnika",
    "Kompyuter injiniringi (Kompyuter injiniringi)": "IT va Texnika",
    "Iqtisodiyot": "IT va Texnika",
    "Iqtisodiyot (tarmoqlar va sohalar bo'yicha)": "IT va Texnika",
    "Moliya va moliyaviy texnologiyalar": "IT va Texnika",
    
    # Tibbiyot fakulteti
    "Davolash ishi": "Tibbiyot",
    "Pediatriya ishi": "Tibbiyot",
    "Stomatologiya": "Tibbiyot",
    "Farmatsiya": "Tibbiyot",
    "Farmatsiya (turlari bo'yicha)": "Tibbiyot",
    
    # Ijtimoiy-gumanitar fakulteti
    "Boshlang\u0027ich ta\u0027lim": "Ijtimoiy-gumanitar",
    "Boshlangʻich taʼlim": "Ijtimoiy-gumanitar",
    "Maktabgacha ta'lim": "Ijtimoiy-gumanitar",
    "Filologiya va tillarni o'qitish (ingliz tili)": "Ijtimoiy-gumanitar",
    "Filologiya va tillarni o'qitish (o'zbek tili)": "Ijtimoiy-gumanitar",
    "Filologiya va tillarni o'qitish (rus tili)": "Ijtimoiy-gumanitar",
    "Lingvistika (ingliz tili)": "Ijtimoiy-gumanitar",
    "Psixologiya": "Ijtimoiy-gumanitar",
    "Psixologiya (faoliyat turlari bo'yicha)": "Ijtimoiy-gumanitar",
    "Tarix": "Ijtimoiy-gumanitar",
    "Tarix (mamlakatlar va yo'nalishlar bo'yicha)": "Ijtimoiy-gumanitar",
}

FACULTY_ORDER = ["IT va Texnika", "Tibbiyot", "Ijtimoiy-gumanitar"]


def get_super_faculty(direction: str) -> str:
    """Get super-faculty for a direction. Falls back to 'Boshqa' if not found."""
    if not direction:
        return "Boshqa"
    # Exact match first
    if direction in FACULTY_MAPPING:
        return FACULTY_MAPPING[direction]
    # Try case-insensitive match
    for key, val in FACULTY_MAPPING.items():
        if key.lower() == direction.lower():
            return val
    # Try partial match
    direction_lower = direction.lower()
    if "kompyuter" in direction_lower or "iqtisod" in direction_lower or "moliya" in direction_lower:
        return "IT va Texnika"
    if "davolash" in direction_lower or "pediatr" in direction_lower or "stomat" in direction_lower or "farma" in direction_lower:
        return "Tibbiyot"
    return "Ijtimoiy-gumanitar"


@router.get("/faculties-tree")
async def academic_faculties_tree(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """
    Get hierarchical faculty → direction → groups tree from kontingent data.
    Groups come from student/group database (kontingent), not from schedule.
    """
    # Get all active groups with student counts
    result = await db.execute(
        select(
            Group.id,
            Group.name,
            Group.faculty,
            Group.course_year,
            func.count(Student.id).label('students_count')
        )
        .outerjoin(Student, and_(Student.group_id == Group.id, Student.is_active == True))
        .where(Group.is_active == True)
        .group_by(Group.id, Group.name, Group.faculty, Group.course_year)
        .order_by(Group.faculty, Group.name)
    )
    rows = result.all()
    
    # Build tree: faculty → directions → groups
    tree = {}
    for row in rows:
        direction = row.faculty or "Boshqa"
        super_faculty = get_super_faculty(direction)
        
        if super_faculty not in tree:
            tree[super_faculty] = {}
        if direction not in tree[super_faculty]:
            tree[super_faculty][direction] = []
        
        tree[super_faculty][direction].append({
            "id": row.id,
            "name": row.name,
            "course_year": row.course_year,
            "students_count": row.students_count or 0,
        })
    
    # Format as ordered list
    faculties = []
    for faculty_name in FACULTY_ORDER:
        if faculty_name in tree:
            directions = []
            for dir_name, groups in sorted(tree[faculty_name].items()):
                directions.append({
                    "name": dir_name,
                    "groups": sorted(groups, key=lambda g: g["name"]),
                    "groups_count": len(groups),
                    "students_count": sum(g["students_count"] for g in groups),
                })
            faculties.append({
                "name": faculty_name,
                "directions": directions,
                "directions_count": len(directions),
                "groups_count": sum(d["groups_count"] for d in directions),
                "students_count": sum(d["students_count"] for d in directions),
            })
    
    # Add 'Boshqa' if any unmatched
    if "Boshqa" in tree:
        directions = []
        for dir_name, groups in sorted(tree["Boshqa"].items()):
            directions.append({
                "name": dir_name,
                "groups": sorted(groups, key=lambda g: g["name"]),
                "groups_count": len(groups),
                "students_count": sum(g["students_count"] for g in groups),
            })
        faculties.append({
            "name": "Boshqa",
            "directions": directions,
            "directions_count": len(directions),
            "groups_count": sum(d["groups_count"] for d in directions),
            "students_count": sum(d["students_count"] for d in directions),
        })
    
    return {
        "faculties": faculties,
        "total_faculties": len(faculties),
        "total_directions": sum(f["directions_count"] for f in faculties),
        "total_groups": sum(f["groups_count"] for f in faculties),
    }


# ============================================
# AI GENERATION
# ============================================

@router.post("/ai/generate")
async def academic_ai_generate(
    data: AIGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Generate schedule using Claude AI."""
    
    # Get groups
    groups_result = await db.execute(
        select(Group).where(Group.id.in_(data.group_ids))
    )
    groups = [{"id": g.id, "name": g.name} for g in groups_result.scalars().all()]
    
    if not groups:
        raise HTTPException(status_code=400, detail="Guruhlar topilmadi")
    
    # Get teachers
    teachers_result = await db.execute(
        select(User).where(and_(User.role == UserRole.TEACHER, User.is_active == True))
    )
    teachers = [{"id": t.id, "name": t.name} for t in teachers_result.scalars().all()]
    
    # Generate
    subjects_for_ai = [
        {"name": s.name, "type": s.type, "hours_per_week": s.hours_per_week}
        for s in data.subjects
    ]
    result = await generate_schedule_with_ai(
        groups=groups,
        subjects=subjects_for_ai,
        teachers=teachers,
        rooms=data.rooms or [],
        constraints=data.constraints,
        language=data.language
    )
    
    return result


@router.post("/ai/optimize")
async def academic_ai_optimize(
    data: AIOptimizeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Optimize existing schedule using AI."""
    
    query = select(Schedule).options(joinedload(Schedule.group)).where(Schedule.is_active == True)
    if data.group_id:
        query = query.where(Schedule.group_id == data.group_id)
    
    result = await db.execute(query)
    schedules = result.unique().scalars().all()
    
    current = [schedule_to_dict(s) for s in schedules]
    
    ai_result = await ai_optimize_schedule(
        current_schedules=current,
        problem_description=data.problem,
        language=data.language
    )
    
    return ai_result


# ============================================
# TEACHER WORKLOAD IMPORT
# ============================================

class WorkloadImportItem(BaseModel):
    teacher_name: str
    department: Optional[str] = None
    teacher_type: Optional[str] = None
    day_of_week: str  # monday, tuesday, etc.
    day_name_uz: Optional[str] = None
    lesson_number: int
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    groups: str
    is_busy: bool = False

class WorkloadImportRequest(BaseModel):
    items: List[WorkloadImportItem]
    clear_existing: bool = True  # Clear old data before import

@router.post("/workload/import")
async def academic_import_workload(
    data: WorkloadImportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Import teacher workload data (from Excel parsing)."""
    
    if data.clear_existing:
        await db.execute(delete(TeacherWorkload))
        await db.flush()
    
    imported = 0
    for item in data.items:
        workload = TeacherWorkload(
            teacher_name=item.teacher_name,
            department=item.department,
            teacher_type=item.teacher_type,
            day_of_week=item.day_of_week,
            day_name_uz=item.day_name_uz,
            lesson_number=item.lesson_number,
            start_time=item.start_time,
            end_time=item.end_time,
            groups=item.groups,
            is_busy=item.is_busy,
            source_file="excel_import",
            is_active=True
        )
        db.add(workload)
        imported += 1
    
    await db.commit()
    
    return {
        "success": True,
        "message": f"{imported} ta o'qituvchi bandligi yuklandi",
        "total_imported": imported
    }

@router.get("/workload/stats")
async def academic_workload_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Get workload import stats."""
    total = await db.execute(
        select(func.count(TeacherWorkload.id)).where(TeacherWorkload.is_active == True)
    )
    teachers_count = await db.execute(
        select(func.count(distinct(TeacherWorkload.teacher_name))).where(TeacherWorkload.is_active == True)
    )
    departments_count = await db.execute(
        select(func.count(distinct(TeacherWorkload.department))).where(
            and_(TeacherWorkload.is_active == True, TeacherWorkload.department.isnot(None))
        )
    )
    
    return {
        "total_entries": total.scalar() or 0,
        "total_teachers": teachers_count.scalar() or 0,
        "total_departments": departments_count.scalar() or 0,
    }


# ============================================
# EXCEL FILE UPLOAD - WORKLOAD IMPORT
# ============================================

LESSON_TIMES = {
    1: ("08:30", "09:50"),
    2: ("10:00", "11:20"),
    3: ("11:30", "12:50"),
    4: ("13:30", "14:50"),
    5: ("15:00", "16:20"),
    6: ("16:30", "17:50"),
    7: ("18:00", "19:20"),
}

DAY_MAP_EXCEL = {
    "dushanba": "monday",
    "seshanba": "tuesday",
    "chorshanba": "wednesday",
    "payshanba": "thursday",
    "juma": "friday",
    "shanba": "saturday",
}


@router.post("/workload/upload-excel")
async def upload_workload_excel(
    file: UploadFile = File(...),
    clear_existing: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Upload Excel file and parse teacher workload sheet."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Faqat Excel (.xlsx, .xls) fayllari qabul qilinadi")

    try:
        import openpyxl
        import io

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        # Find workload sheet
        sheet = None
        for name in wb.sheetnames:
            if "bandlig" in name.lower() or "band" in name.lower():
                sheet = wb[name]
                break

        if sheet is None:
            return {
                "success": False,
                "message": "'Oqituvchi bandligi' varag'i topilmadi",
                "available_sheets": wb.sheetnames
            }

        # Parse sheet
        items = []
        current_teacher = None
        current_department = None
        current_type = None

        for row in sheet.iter_rows(min_row=1, values_only=False):
            vals = [cell.value for cell in row]
            if all(v is None for v in vals):
                continue

            first_val = str(vals[0]).strip() if vals[0] else ""

            # Department header
            if any(word in first_val.lower() for word in ["kafedra", "bo'lim"]):
                current_department = first_val
                continue

            # Check for day of week
            day_uz = None
            for dname in DAY_MAP_EXCEL.keys():
                if dname in first_val.lower():
                    day_uz = dname
                    break

            if day_uz and current_teacher:
                day_en = DAY_MAP_EXCEL[day_uz]
                for col_idx in range(1, len(vals)):
                    cell_val = vals[col_idx]
                    if cell_val is None:
                        continue
                    cell_str = str(cell_val).strip()
                    if not cell_str:
                        continue
                    lesson_num = col_idx
                    times = LESSON_TIMES.get(lesson_num, ("", ""))
                    is_busy = cell_str.upper() in ("BAND", "X")
                    items.append({
                        "teacher_name": current_teacher,
                        "department": current_department,
                        "teacher_type": current_type,
                        "day_of_week": day_en,
                        "day_name_uz": day_uz.capitalize(),
                        "lesson_number": lesson_num,
                        "start_time": times[0],
                        "end_time": times[1],
                        "groups": "BAND" if is_busy else cell_str,
                        "is_busy": is_busy,
                    })
                continue

            # Teacher name (2-6 words, no digits, not a day)
            if not day_uz and first_val and len(first_val) > 3:
                words = first_val.split()
                if 2 <= len(words) <= 6 and not any(c.isdigit() for c in first_val):
                    teacher_type = None
                    for ttype in ["professor", "dotsent", "katta o'qituvchi", "o'qituvchi", "assistent"]:
                        if ttype in first_val.lower():
                            teacher_type = ttype
                            break
                    current_teacher = first_val
                    current_type = teacher_type
                    continue

        if not items:
            return {
                "success": False,
                "message": "Varaqdan ma'lumot topilmadi",
                "sheet_name": sheet.title,
            }

        # Import to DB
        if clear_existing:
            await db.execute(delete(TeacherWorkload))
            await db.flush()

        imported = 0
        for item in items:
            workload = TeacherWorkload(
                teacher_name=item["teacher_name"],
                department=item["department"],
                teacher_type=item["teacher_type"],
                day_of_week=item["day_of_week"],
                day_name_uz=item["day_name_uz"],
                lesson_number=item["lesson_number"],
                start_time=item["start_time"],
                end_time=item["end_time"],
                groups=item["groups"],
                is_busy=item["is_busy"],
                source_file=file.filename,
                is_active=True
            )
            db.add(workload)
            imported += 1

        await db.commit()

        teachers = set(i["teacher_name"] for i in items)
        depts = set(i["department"] for i in items if i["department"])

        return {
            "success": True,
            "message": f"{imported} ta yozuv import qilindi",
            "total_imported": imported,
            "total_teachers": len(teachers),
            "total_departments": len(depts),
            "source_file": file.filename,
            "sheet_name": sheet.title
        }

    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kutubxonasi o'rnatilmagan")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel xatolik: {str(e)}")


# ============================================
# SCHEDULE EXPORT - ALL GROUPS TO EXCEL
# ============================================

@router.get("/schedules/export")
async def academic_export_all_schedules(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_academic)
):
    """Export all groups' schedules to a single Excel file with one sheet per group."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Get all active groups that have schedules
    groups_result = await db.execute(
        select(Group).where(Group.is_active == True).order_by(Group.name)
    )
    all_groups = groups_result.scalars().all()

    # Get all active schedules
    schedules_result = await db.execute(
        select(Schedule)
        .where(Schedule.is_active == True)
        .order_by(Schedule.day_of_week, Schedule.lesson_number)
    )
    all_schedules = schedules_result.scalars().all()

    # Group schedules by group_id
    from collections import defaultdict
    schedules_by_group = defaultdict(list)
    for s in all_schedules:
        schedules_by_group[s.group_id].append(s)

    day_names = {
        "monday": "Dushanba", "tuesday": "Seshanba", "wednesday": "Chorshanba",
        "thursday": "Payshanba", "friday": "Juma", "saturday": "Shanba",
    }
    day_order = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]

    para_times = {
        1: ("08:30", "09:50"), 2: ("10:00", "11:20"), 3: ("11:30", "12:50"),
        4: ("13:30", "14:50"), 5: ("15:00", "16:20"), 6: ("16:30", "17:50"),
        7: ("18:00", "19:20"),
    }

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    day_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    day_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    groups_exported = 0

    for group in all_groups:
        group_schedules = schedules_by_group.get(group.id, [])
        if not group_schedules:
            continue

        # Sheet name max 31 chars
        sheet_name = (group.name or f"Group_{group.id}")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(1, 1, f"Dars jadvali: {group.name}")
        title_cell.font = Font(bold=True, size=14, color="1E40AF")
        title_cell.alignment = center

        # Headers
        headers = ["Kun", "Para", "Vaqt", "Fan", "O'qituvchi", "Xona", "Hafta"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(3, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        row = 4
        for day_val in day_order:
            day_lessons = [s for s in group_schedules if (s.day_of_week.value if s.day_of_week else "") == day_val]
            if not day_lessons:
                continue

            day_lessons.sort(key=lambda s: s.lesson_number or 0)

            for s in day_lessons:
                dv = s.day_of_week.value if s.day_of_week else ""
                ln = s.lesson_number or 0
                st = s.start_time.strftime("%H:%M") if s.start_time else (para_times.get(ln, ("", ""))[0])
                et = s.end_time.strftime("%H:%M") if s.end_time else (para_times.get(ln, ("", ""))[1])

                ws.cell(row, 1, day_names.get(dv, dv)).font = day_font
                ws.cell(row, 1).fill = day_fill
                ws.cell(row, 2, ln)
                ws.cell(row, 3, f"{st} - {et}")
                ws.cell(row, 4, s.subject or "")
                ws.cell(row, 5, s.teacher_name or "")
                ws.cell(row, 6, s.room or "")
                wt = s.week_type.value if s.week_type else "all"
                wt_label = {"all": "Har hafta", "odd": "Toq hafta", "even": "Juft hafta"}.get(wt, wt)
                ws.cell(row, 7, wt_label)

                for col in range(1, 8):
                    ws.cell(row, col).border = thin_border
                    ws.cell(row, col).alignment = center

                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12

        groups_exported += 1

    # If no schedules, create a summary sheet
    if groups_exported == 0:
        ws = wb.create_sheet(title="Bo'sh")
        ws.cell(1, 1, "Hech qanday guruh jadvali topilmadi")

    # Also create a summary sheet
    summary = wb.create_sheet(title="Umumiy", index=0)
    summary.cell(1, 1, "Dars jadvali — barcha guruhlar").font = Font(bold=True, size=14)
    summary.cell(3, 1, "Guruh").font = Font(bold=True)
    summary.cell(3, 2, "Darslar soni").font = Font(bold=True)
    summary.cell(3, 3, "Fakultet").font = Font(bold=True)

    sr = 4
    for group in all_groups:
        cnt = len(schedules_by_group.get(group.id, []))
        if cnt == 0:
            continue
        summary.cell(sr, 1, group.name)
        summary.cell(sr, 2, cnt)
        summary.cell(sr, 3, group.faculty or "")
        sr += 1

    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 14
    summary.column_dimensions['C'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=dars_jadvali_barcha_guruhlar.xlsx"
        }
    )
