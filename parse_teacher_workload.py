import openpyxl
import glob
import os
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

files = glob.glob(os.path.join(r"c:\Users\user\Desktop\unicontrol", "1-2*dars jadvali*.xlsx"))
fname = files[0]
wb = openpyxl.load_workbook(fname)
ws = wb["Oqituvchi bandligi"]

merged_map = {}
for mc in ws.merged_cells.ranges:
    for row in range(mc.min_row, mc.max_row + 1):
        for col in range(mc.min_col, mc.max_col + 1):
            merged_map[(row, col)] = ws.cell(mc.min_row, mc.min_col).value

def get_val(row, col):
    v = ws.cell(row, col).value
    if v is not None:
        return v
    return merged_map.get((row, col), None)

# Parse departments from row 1
dept_ranges = {}
for col in range(4, ws.max_column + 1):
    v = get_val(1, col)
    if v:
        dept_name = str(v).strip()
        if dept_name not in dept_ranges:
            dept_ranges[dept_name] = {"start": col, "end": col}
        dept_ranges[dept_name]["end"] = col

# Parse teachers from row 2
teachers = {}
for col in range(4, ws.max_column + 1):
    v = get_val(2, col)
    if v:
        teacher_type = get_val(5, col) or ""
        # Find department
        dept = ""
        for d_name, d_range in dept_ranges.items():
            if d_range["start"] <= col <= d_range["end"]:
                dept = d_name
                break
        teachers[col] = {
            "name": str(v).strip(),
            "type": str(teacher_type).strip(),
            "department": dept,
            "col": col
        }

# Parse schedule - day/lesson/time mapping
DAYS_MAP = {
    "Dushanba": "monday",
    "Seshanba": "tuesday", 
    "Chorshanba": "wednesday",
    "Payshanba": "thursday",
    "Juma": "friday",
    "Juma ": "friday",
    "Shanba": "saturday"
}

TIME_MAP = {
    1: {"start": "08:30", "end": "09:50"},
    2: {"start": "10:00", "end": "11:20"},
    3: {"start": "12:00", "end": "13:20"},
    4: {"start": "13:30", "end": "14:50"},
    5: {"start": "15:00", "end": "16:20"},
    6: {"start": "16:30", "end": "17:50"},
}

schedule_entries = []
current_day = None
current_day_en = None

for row in range(6, ws.max_row + 1):
    day_val = get_val(row, 1)
    lesson_num = get_val(row, 2)
    
    if day_val:
        current_day = str(day_val).strip()
        current_day_en = DAYS_MAP.get(current_day, current_day.lower())
    
    if not lesson_num or not current_day:
        continue
    
    ln = int(float(lesson_num))
    time_info = TIME_MAP.get(ln, {"start": "", "end": ""})
    
    for col, teacher in teachers.items():
        group_val = get_val(row, col)
        if group_val and str(group_val).strip():
            gv = str(group_val).strip()
            if gv == "BAND":
                schedule_entries.append({
                    "day": current_day,
                    "day_en": current_day_en,
                    "lesson_number": ln,
                    "start_time": time_info["start"],
                    "end_time": time_info["end"],
                    "teacher_name": teacher["name"],
                    "teacher_type": teacher["type"],
                    "department": teacher["department"],
                    "groups": "BAND",
                    "is_busy": True
                })
            else:
                schedule_entries.append({
                    "day": current_day,
                    "day_en": current_day_en,
                    "lesson_number": ln,
                    "start_time": time_info["start"],
                    "end_time": time_info["end"],
                    "teacher_name": teacher["name"],
                    "teacher_type": teacher["type"],
                    "department": teacher["department"],
                    "groups": gv,
                    "is_busy": False
                })

# Output summary
result = {
    "departments": list(dept_ranges.keys()),
    "total_teachers": len(teachers),
    "total_entries": len(schedule_entries),
    "teachers": [{"name": t["name"], "type": t["type"], "department": t["department"]} for t in teachers.values()],
    "schedule": schedule_entries
}

with open(r"c:\Users\user\Desktop\unicontrol\teacher_workload.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Departments: {len(dept_ranges)}")
for d in dept_ranges:
    print(f"  - {d}")
print(f"Teachers: {len(teachers)}")
print(f"Schedule entries: {len(schedule_entries)}")
print(f"Saved to teacher_workload.json")
